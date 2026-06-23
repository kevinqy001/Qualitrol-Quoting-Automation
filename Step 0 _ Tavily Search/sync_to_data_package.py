"""Sync the Tavily-researched product catalog into the master data package.

Step 0 writes its results to
    outputs/_product_catalog/Qualitrol_Product_Catalog.xlsx
(sheets 06 / 07 / 08).

This script promotes those rows into the matching master file
    Qualitrol_BOQ_Matching_Data_Package.xlsx
by replacing the TBD placeholder rows in sheets 06 / 07 / 08.
A timestamped backup of the master is created before any write.

Usage (run from the project root):
    python "Step 0 _ Tavily Search/sync_to_data_package.py"

Options:
    --dry-run   Preview what would change without writing anything.
    --force     Skip the confirmation prompt.
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from qualitrol_core.config import DATA_PACKAGE_PATH  # noqa: E402

CATALOG_PATH = ROOT / "outputs" / "_product_catalog" / "Qualitrol_Product_Catalog.xlsx"
BACKUP_DIR = ROOT / "outputs" / "_product_catalog" / "backups"

# Sheets to sync: sheet_name -> first column of the header row
SYNC_SHEETS = {
    "06_Product_Family_Master": "Product Family ID",
    "07_Product_Master_Template": "Product ID",
    "08_Product_Parameter_Template": "Product ID",
}


def _find_header_row(ws, first_col_value: str) -> int:
    """Return the 1-based row index of the header row in a master sheet."""
    target = first_col_value.strip().lower()
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value is not None and str(cell.value).strip().lower() == target:
            return cell.row
    raise ValueError(
        f"Header row starting with '{first_col_value}' not found in sheet '{ws.title}'"
    )


def _catalog_data(cat_wb, sheet_name: str) -> tuple[list[str], list[list]]:
    """Read the catalog sheet; return (header_cols, data_rows).

    The catalog sheet has a single header row (row 0) followed by data rows.
    """
    ws = cat_wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [
        list(r)
        for r in rows[1:]
        if any(c is not None and str(c).strip() not in ("", "None") for c in r)
    ]
    return headers, data


def _align_rows_to_master_header(
    cat_headers: list[str],
    cat_rows: list[list],
    master_headers: list[str],
) -> list[list]:
    """Re-order catalog data columns to match the master header column order.

    Any catalog column not found in the master is dropped.
    Any master column not in the catalog is written as None.
    """
    # Build a mapping: master_col_idx -> catalog_col_idx (or None if absent)
    col_map: list[int | None] = []
    for mh in master_headers:
        try:
            col_map.append(cat_headers.index(mh))
        except ValueError:
            col_map.append(None)

    aligned = []
    for row in cat_rows:
        new_row = []
        for cat_idx in col_map:
            if cat_idx is None:
                new_row.append(None)
            else:
                new_row.append(row[cat_idx] if cat_idx < len(row) else None)
        aligned.append(new_row)
    return aligned


def sync(dry_run: bool = False, force: bool = False) -> None:
    import openpyxl

    if not CATALOG_PATH.exists():
        print(f"[ERROR] Catalog not found:\n  {CATALOG_PATH}")
        print(
            "\nRun Step 0 first to generate the catalog:\n"
            '  python "Step 0 _ Tavily Search/run.py"'
        )
        sys.exit(1)

    if not DATA_PACKAGE_PATH.exists():
        print(f"[ERROR] Master data package not found:\n  {DATA_PACKAGE_PATH}")
        sys.exit(1)

    cat_wb = openpyxl.load_workbook(str(CATALOG_PATH), read_only=True, data_only=True)
    master_wb_ro = openpyxl.load_workbook(
        str(DATA_PACKAGE_PATH), read_only=True, data_only=True
    )

    # ── Preview ────────────────────────────────────────────────────────────── #
    print("=" * 70)
    print("SYNC PREVIEW — catalog  →  master data package")
    print("=" * 70)
    # changes: sheet_name -> aligned rows ready to write into master
    changes: dict[str, list[list]] = {}
    for sheet_name, header_col in SYNC_SHEETS.items():
        if sheet_name not in cat_wb.sheetnames:
            print(f"  [SKIP] {sheet_name}: not in catalog")
            continue
        cat_headers, cat_rows = _catalog_data(cat_wb, sheet_name)

        # Read the master's header row to know its column order
        master_header: list[str] = []
        if sheet_name in master_wb_ro.sheetnames:
            for row in master_wb_ro[sheet_name].iter_rows():
                if row[0].value and str(row[0].value).strip().lower() == header_col.lower():
                    master_header = [
                        str(c.value).strip() if c.value is not None else ""
                        for c in row
                    ]
                    break

        if master_header:
            aligned = _align_rows_to_master_header(cat_headers, cat_rows, master_header)
        else:
            aligned = cat_rows  # fallback: positional (original behaviour)

        changes[sheet_name] = aligned
        first_models = [r[1] for r in aligned[:3] if len(r) > 1 and r[1]]
        print(f"  {sheet_name}: {len(aligned)} rows  "
              f"(e.g. {', '.join(str(m) for m in first_models[:3])})")

    cat_wb.close()
    master_wb_ro.close()

    if dry_run:
        print("\n[dry-run] No files modified.")
        return

    # ── Confirmation ─────────────────────────────────────────────────────── #
    if not force:
        answer = input(
            "\nProceed? The master data package will be backed up then updated. [y/N] "
        )
        if answer.strip().lower() not in ("y", "yes"):
            print("Aborted.")
            return

    # ── Backup ───────────────────────────────────────────────────────────── #
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"Qualitrol_BOQ_Matching_Data_Package__{ts}.xlsx"
    shutil.copy2(DATA_PACKAGE_PATH, backup)
    print(f"\nBacked up → {backup.name}")

    # ── Write ─────────────────────────────────────────────────────────────── #
    # Try loading the master directly; if it's locked (open in Excel), fall back
    # to loading from the backup we just created and saving to a separate
    # "_patched" file that the user can manually rename.
    try:
        master_wb = openpyxl.load_workbook(str(DATA_PACKAGE_PATH), data_only=True)
        output_path = DATA_PACKAGE_PATH
    except PermissionError:
        print(
            "\n[WARNING] Master file is currently open (likely in Excel).\n"
            "          Changes will be written to a patched copy instead.\n"
            "          Close Excel, then replace the master with the patched file."
        )
        master_wb = openpyxl.load_workbook(str(backup), data_only=True)
        output_path = BACKUP_DIR / f"Qualitrol_BOQ_Matching_Data_Package__patched_{ts}.xlsx"

    for sheet_name, new_rows in changes.items():
        if sheet_name not in master_wb.sheetnames:
            print(f"  [SKIP] {sheet_name}: not in master (will not modify)")
            continue

        master_ws = master_wb[sheet_name]
        header_col = SYNC_SHEETS[sheet_name]

        try:
            header_row_idx = _find_header_row(master_ws, header_col)
        except ValueError as exc:
            print(f"  [SKIP] {sheet_name}: {exc}")
            continue

        # Delete every row after the header.
        last_row = master_ws.max_row
        rows_to_delete = last_row - header_row_idx
        if rows_to_delete > 0:
            master_ws.delete_rows(header_row_idx + 1, rows_to_delete)

        # Append catalog data rows.
        for row_data in new_rows:
            master_ws.append(row_data)

        print(f"  [OK]   {sheet_name}: replaced TBD stubs → {len(new_rows)} rows")

    master_wb.save(str(output_path))
    master_wb.close()

    if output_path == DATA_PACKAGE_PATH:
        print(f"\nSaved → {DATA_PACKAGE_PATH.name}")
        print(
            "\nNext step: restart the web server to reload the data package cache:\n"
            "  python app.py"
        )
    else:
        print(f"\nPatched file saved → {output_path.name}")
        print(
            f"\nTo apply:\n"
            f"  1. Close Qualitrol_BOQ_Matching_Data_Package.xlsx in Excel\n"
            f"  2. Replace it with the patched file:\n"
            f"       {output_path}\n"
            f"  3. Then restart the web server: python app.py\n"
            f"\nOR re-run this script after closing Excel and it will update directly."
        )


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    force = "--force" in sys.argv
    sync(dry_run=dry, force=force)
