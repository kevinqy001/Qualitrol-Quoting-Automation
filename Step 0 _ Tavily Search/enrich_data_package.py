# -*- coding: utf-8 -*-
"""
enrich_data_package.py
======================
Enrich Qualitrol_BOQ_Matching_Data_Package.xlsx with:
 - New scenarios  (Sheet 03)
 - New metrics    (Sheet 04)
 - New synonyms   (Sheet 05)
 - New product models  (Sheet 07)
 - New product params  (Sheet 08)

Only APPENDS new rows; never overwrites existing verified content.
Run from the repo root:
    python "Step 0 _ Tavily Search/enrich_data_package.py"
"""

from pathlib import Path
import shutil, datetime, sys, io
import openpyxl
from openpyxl.styles import Font

# ── Paths ──────────────────────────────────────────────────────────────────
REPO   = Path(__file__).resolve().parent.parent
MASTER = REPO / "Qualitrol_BOQ_Matching_Data_Package.xlsx"
BACKUP = REPO / f"Qualitrol_BOQ_Matching_Data_Package_BACKUP_{datetime.date.today()}.xlsx"

# ── Helpers ────────────────────────────────────────────────────────────────
def load_wb():
    tmp = Path(f"C:/Windows/Temp/qboq_enrich_{datetime.datetime.now().strftime('%H%M%S')}.xlsx")
    shutil.copy2(MASTER, tmp)
    return openpyxl.load_workbook(tmp), tmp

def get_header_row(ws):
    """Return (row_index, header_dict) for first row with >=3 non-empty cells."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if sum(1 for c in row if c) >= 3:
            return i, {str(c).strip(): j for j, c in enumerate(row) if c}
    return 1, {}

def existing_values(ws, col_idx):
    """Return set of lower-cased values in col_idx (0-based) across all rows."""
    vals = set()
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) > col_idx and row[col_idx]:
            vals.add(str(row[col_idx]).strip().lower())
    return vals

def append_row(ws, values: list):
    ws.append(values)

# ── NEW CONTENT DEFINITIONS ────────────────────────────────────────────────

# ── Sheet 03: New Scenarios ────────────────────────────────────────────────
NEW_SCENARIOS = [
    # Scenario ID | Category | Application Scenario | Asset Type | Typical Metrics
    # | Common Keywords | Related Product Families | Qty Basis | Drawing Dep | Req Fields | Notes
    [
        "FMS_001",
        "Grid / Disturbance",
        "Centralized fault management system and DFR data platform",
        "Multi-DFR substation / substation network server",
        "Server count; iQ+ license count; managed device count; communication network",
        "FMS; fault management system; iQ+; iQ plus; master station; CASHEL FMS; "
        "centralized recording; central server; substation data server; DFR master station; "
        "iQ master station; data management platform; rackmount server; network server",
        "PF_DFR; PF_TWS; PF_PMU; PF_SOFTWARE",
        "1 server system per site; additional client workstation licenses per operator",
        "Network device inventory and site topology helpful",
        "server_count; managed_device_count; site_count; client_count",
        "Always paired with 2+ IDM+/TWS deployments. iQ+ Master Station is the standard software.",
    ],
    [
        "GIS_ARC_001",
        "GIS",
        "GIS arc flash detection and localization",
        "Gas-insulated switchgear / GIS compartment",
        "Arc event count; arc location; gas compartment; UHF signal level; flashover",
        "arc detection; arc flash; arc localization; GIS arc; flashover; GIS flashover; "
        "arc discharge; flashover detection; arc fault detection; arc event",
        "PF_GIS_PD",
        "No additional hardware; feature of PDMG-RH Gen3 (uses existing UHF sensors)",
        "GIS layout and compartment list required for localization",
        "gis_bay_count; arc_detection_option; compartment_list",
        "Arc Detection is an add-on feature of PDMG-RH Gen3. No extra hardware needed. "
        "Only quote if PDMG-RH Gen3 is already in scope.",
    ],
    [
        "WAMS_001",
        "Grid / Synchrophasor",
        "Wide area monitoring system for multi-site grid stability",
        "Multi-substation transmission grid / phasor data concentrator",
        "PMU count; phasor data concentrator count; reporting rate; GPS latency; site count",
        "WAMS; wide area monitoring; wide area measurement system; PDC; "
        "phasor data concentrator; real-time oscillation monitoring; power swing monitoring; "
        "inter-area oscillation; grid stability monitoring; frequency monitoring; voltage angle",
        "PF_PMU; PF_SOFTWARE",
        "1 PMU per monitored grid measurement point; 1 PDC per substation cluster",
        "Grid topology and SLD required; multi-site coordination",
        "pmu_count; site_count; grid_segment; reporting_rate",
        "Typically integrated with GCCIA or national SCADA. PMU_001 is per-device; "
        "WAMS_001 covers the system-wide wide-area network use case.",
    ],
]

# ── Sheet 04: New Metrics ────────────────────────────────────────────────
# ID | Name | Synonyms | Unit | Data Type | Applies To | Used For | Example | Required | Notes
NEW_METRICS = [
    [
        "MET_FEEDER_COUNT",
        "Feeder / Circuit Count",
        "feeder count; feeder; number of feeders; bay count; circuit count; number of circuits; "
        "feeders per panel; protected feeders; monitored feeders",
        "count",
        "integer",
        "DFR; PQ; PMU; TWS",
        "DFR model selection (determines analog channel count needed); TWS line module count",
        "6 feeders; 12 bays",
        "Yes (for DFR/TWS sizing)",
        "1 feeder = typically 3 current + 1-3 voltage channels for full DFR coverage.",
    ],
    [
        "MET_LINE_COUNT",
        "Transmission Line Count",
        "line count; number of lines; transmission lines; monitored lines; line ends; circuits",
        "count",
        "integer",
        "TWS; DFR",
        "TWS FL module selection (FL-8 = 2/4/6/8 lines; FL-1 = 1 line)",
        "4 transmission lines; 2 line ends",
        "Yes (for TWS)",
        "TWS FL-8 starts with 2-line module, expandable to 8 in steps of 2.",
    ],
    [
        "MET_ANALOG_CHANNELS",
        "Total Analog Channel Count",
        "analog channels; AI count; analog inputs; channel count; total channels; "
        "number of analog; A/D channels; recording channels",
        "count",
        "integer",
        "DFR; PQ; PMU; INFORMA PMD-A",
        "IDM+ model/variant selection (9, 18, or 36 analog channels)",
        "9 analog channels; 18 AI; 36A",
        "Yes (for DFR model selection)",
        "IDM+ 9=9A/32D, IDM+ 18=18A/64D, IDM+ 36=36A/128D. "
        "Rule of thumb: 3 current + 3 voltage per feeder = 6 channels/feeder.",
    ],
    [
        "MET_DFR_FUNCTIONS",
        "Required DFR / Recorder Functions",
        "DFR functions; recording functions; monitoring options; PMU option; "
        "TWS option; Class A PQ option; SOE option; multifunctional recorder",
        "text",
        "controlled text (multi-select)",
        "DFR; PMU; TWS",
        "Determines which optional functions to order with IDM+: PMU, TWS card, Class A PQ",
        "DFR + PMU; DFR + TWS + PMU; full functions",
        "Yes (where stated)",
        "IDM+ base = DFR+DDR. Options: +PMU (IEEE C37.118), +TWS card (±60m FL), "
        "+Class A PQ (IEC 61000-4-30 Class A).",
    ],
    [
        "MET_GPS_SYNC_ACCURACY",
        "GPS Time Sync Accuracy",
        "GPS accuracy; time tagging accuracy; timestamp accuracy; GPS time reference; "
        "time synchronization accuracy; 1PPS accuracy; IRIG-B accuracy",
        "ns",
        "number",
        "TWS; PMU; DFR",
        "Verify GPS accuracy meets double-ended TWS requirement (100 ns) and PMU C37.118",
        "100 ns GPS; 1 µs IRIG-B",
        "Yes (for TWS double-ended)",
        "TWS double-ended requires GPS 100 ns. IDM+ GPS input provides 100 ns tagging.",
    ],
]

# ── Sheet 05: New Synonyms ────────────────────────────────────────────────
# Raw Term | Mapped Scenario ID | Mapped Metric ID | Mapped Standard Meaning | Asset Context | Priority
NEW_SYNONYMS = [
    # --- DFR_DDR_001 ---
    ["IDM+",                "DFR_DDR_001; PMU_001; FAULT_LOC_001; PQ_CLASSA_001", "MET_SAMPLING_RATE",
     "Qualitrol IDM+ multifunction recorder (DFR/DDR/PMU/TWS/PQ)", "Substation / feeder", "High"],
    ["IDM plus",            "DFR_DDR_001", "",
     "Qualitrol IDM+ multifunction recorder", "Substation / feeder", "High"],
    ["IDM",                 "DFR_DDR_001", "MET_SAMPLING_RATE",
     "Qualitrol IDM/IDM+ fault recorder (legacy or current)", "Substation", "Medium"],
    ["Hathaway DFR",        "DFR_DDR_001", "",
     "Hathaway (Qualitrol) digital fault recorder", "Substation", "High"],
    ["Hathaway IDM",        "DFR_DDR_001; PMU_001", "",
     "Hathaway IDM fault and disturbance recorder with PMU option", "Substation / grid", "High"],
    ["DDR",                 "DFR_DDR_001", "MET_SAMPLING_RATE",
     "Dynamic Disturbance Recorder (long-duration slow scan recording)", "Substation / grid", "High"],
    ["dynamic disturbance recorder", "DFR_DDR_001", "",
     "Long-duration slow scan recording of power system disturbances", "Grid", "High"],
    ["CFR",                 "DFR_DDR_001", "",
     "Continuous Fault Recorder (continuous DFR recording mode)", "Substation", "Medium"],
    ["continuous fault recorder", "DFR_DDR_001", "",
     "IDM+ continuous fault recording function", "Substation", "Medium"],
    ["DFR",                 "DFR_DDR_001", "MET_SAMPLING_RATE",
     "Digital Fault Recorder", "Substation / feeder", "High"],
    ["digital fault recorder", "DFR_DDR_001", "",
     "Digital fault recording device", "Substation", "High"],
    ["DFR panel",           "DFR_DDR_001; FMS_001", "",
     "Fault recording panel housing IDM+ and accessories", "Substation panel", "High"],
    ["FRS panel",           "DFR_DDR_001; FMS_001", "",
     "Fault Recording System panel", "Substation", "High"],
    ["FRS",                 "DFR_DDR_001", "",
     "Fault Recording System (generic term for DFR system)", "Substation", "High"],
    ["fault recording system", "DFR_DDR_001", "",
     "Generic: fault recording system installed at substation", "Substation", "High"],
    ["fault disturbance recorder", "DFR_DDR_001", "",
     "Combined fault and disturbance recorder", "Substation", "High"],
    ["disturbance monitor", "DFR_DDR_001", "",
     "Power system disturbance monitoring", "Substation / grid", "Medium"],
    ["IDM+ 9",              "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 9-channel variant (9A/32D, 3U)", "Substation", "High"],
    ["IDM+ 18",             "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 18-channel variant (18A/64D, 6U)", "Substation", "High"],
    ["IDM+ 36",             "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 36-channel variant (36A/128D, 6U)", "Substation", "High"],
    ["IDM+ 9A",             "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 9 analog channel configuration", "Substation", "High"],
    ["IDM+ 18A",            "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 18 analog channel configuration", "Substation", "High"],
    ["IDM+ 36A",            "DFR_DDR_001", "MET_ANALOG_CHANNELS",
     "IDM+ 36 analog channel configuration", "Substation", "High"],
    ["SOE",                 "DFR_DDR_001", "",
     "Sequence of Events recorder (function of IDM+)", "Substation", "Medium"],
    ["sequence of events",  "DFR_DDR_001", "",
     "SOE recording of digital input state changes", "Substation", "Medium"],
    ["waveform recorder",   "DFR_DDR_001", "",
     "Voltage/current waveform recording capability", "Substation", "Medium"],
    # --- PMU_001 ---
    ["WAMS",                "PMU_001; WAMS_001", "MET_REPORTING_RATE",
     "Wide Area Monitoring System based on synchrophasors", "Transmission grid", "High"],
    ["wide area monitoring", "PMU_001; WAMS_001", "",
     "Wide-area phasor-based grid monitoring", "Multi-site grid", "High"],
    ["wide area measurement", "PMU_001; WAMS_001", "",
     "Wide Area Measurement System (IEEE C37.118)", "Transmission grid", "High"],
    ["PDC",                 "PMU_001; WAMS_001", "",
     "Phasor Data Concentrator (central collection point for PMU data)", "Grid control center", "High"],
    ["phasor data concentrator", "PMU_001; WAMS_001", "",
     "Central server for collecting and processing PMU data streams", "Grid", "High"],
    ["IEEE C37.118",        "PMU_001", "MET_REPORTING_RATE",
     "IEEE standard for synchrophasor measurement", "Grid / PMU", "High"],
    ["C37.118",             "PMU_001", "MET_REPORTING_RATE",
     "IEEE C37.118 PMU compliance standard", "Grid / PMU", "High"],
    ["C37.118 2014",        "PMU_001", "",
     "IEEE C37.118-2014 revised PMU standard", "Grid / PMU", "High"],
    ["synchrophasor streaming", "PMU_001", "",
     "Real-time phasor data streaming compliant with IEEE C37.118", "Grid", "High"],
    ["phasor measurement unit", "PMU_001", "MET_REPORTING_RATE",
     "Synchrophasor measurement device", "Grid", "High"],
    ["PMU panel",           "PMU_001; FMS_001", "",
     "Panel housing PMU-enabled IDM+ devices", "Substation", "High"],
    ["FMS",                 "FMS_001; DFR_DDR_001", "",
     "Fault Monitoring System (iQ+ platform + DFR hardware at substation)", "Multi-site grid", "High"],
    # --- FAULT_LOC_001 ---
    ["FL-8",                "FAULT_LOC_001", "MET_LINE_COUNT",
     "Qualitrol TWS FL-8 traveling wave fault locator (up to 8 lines)", "Transmission line", "High"],
    ["FL-1",                "FAULT_LOC_001", "MET_LINE_COUNT",
     "Qualitrol TWS FL-1 single-line traveling wave fault locator", "Transmission line", "High"],
    ["FL8",                 "FAULT_LOC_001", "",
     "TWS FL-8 fault locator abbreviation", "Transmission line", "High"],
    ["FL1",                 "FAULT_LOC_001", "",
     "TWS FL-1 fault locator abbreviation", "Transmission line", "High"],
    ["TWS FL-8",            "FAULT_LOC_001", "MET_LINE_COUNT",
     "Qualitrol TWS FL-8 product (modular, up to 8 lines)", "Transmission line", "High"],
    ["TWS FL-1",            "FAULT_LOC_001", "",
     "Qualitrol TWS FL-1 product (single-line, distributed substations)", "Transmission line", "High"],
    ["linear coupler",      "FAULT_LOC_001", "",
     "CT secondary linear coupler for TWS traveling wave detection", "Transmission line CT secondary", "High"],
    ["TWS card",            "FAULT_LOC_001", "",
     "TWS option card fitted in IDM+ for integrated fault location", "Substation / IDM+", "High"],
    ["FL2",                 "FAULT_LOC_001", "MET_LINE_COUNT",
     "TWS FL-8 configured for 2-line monitoring (BOQ shorthand)", "Transmission line", "Medium"],
    ["FL-2",                "FAULT_LOC_001", "MET_LINE_COUNT",
     "TWS FL-8 in 2-line configuration (site BOQ notation)", "Transmission line", "Medium"],
    # --- FMS_001 ---
    ["iQ+",                 "FMS_001; SUB_SOFT_001", "",
     "Qualitrol iQ+ Master Station software for DFR/TWS/PMU fleet management", "Substation server", "High"],
    ["iQ plus",             "FMS_001; SUB_SOFT_001", "",
     "iQ+ Master Station (alternate spelling)", "Substation server", "High"],
    ["master station",      "FMS_001; SUB_SOFT_001", "",
     "iQ+ centralized master station software", "Substation server", "High"],
    ["iQ master station",   "FMS_001; SUB_SOFT_001", "",
     "Qualitrol iQ+ Master Station application", "Substation server", "High"],
    ["CASHEL FMS",          "FMS_001", "",
     "CASHEL (Qualitrol brand name for FMS panels/server system)", "Substation FMS", "High"],
    ["CASHEL server",       "FMS_001", "",
     "Rackmount server included in CASHEL FMS system", "Substation server", "High"],
    ["FMS server",          "FMS_001", "",
     "Fault Management System server (iQ+ host)", "Substation server", "High"],
    ["GridAssist",          "FMS_001; SUB_SOFT_001", "",
     "Qualitrol GridAssist substation monitoring platform", "Substation / fleet", "Medium"],
    ["central server",      "FMS_001", "",
     "Central iQ+ server for multi-DFR data collection", "Substation server", "Medium"],
    ["substation data server", "FMS_001", "",
     "Server infrastructure for substation monitoring data management", "Substation server", "Medium"],
    # --- SUB_SOFT_001 ---
    ["SmartSub",            "SUB_SOFT_001", "",
     "Qualitrol SmartSub substation asset performance management platform", "Substation / fleet", "High"],
    # --- GIS_PD_001 ---
    ["PDMG-RH",             "GIS_PD_001", "MET_SENSOR_TYPE",
     "Qualitrol PDMG-RH GIS PD monitoring system (Gen 3)", "GIS", "High"],
    ["PDMG",                "GIS_PD_001", "MET_SENSOR_TYPE",
     "PDMG series partial discharge monitor for GIS", "GIS", "High"],
    ["GIS PDM",             "GIS_PD_001", "",
     "GIS Partial Discharge Monitor (generic abbreviation)", "GIS", "High"],
    ["UHF coupler",         "GIS_PD_001", "MET_SENSOR_TYPE",
     "UHF sensor/coupler installed on GIS for PD detection", "GIS", "High"],
    ["OCU",                 "GIS_PD_001", "",
     "Optical Converter Unit (signal conditioner in PDMG-RH system)", "GIS PDM system", "Medium"],
    ["optical converter unit", "GIS_PD_001", "",
     "OCU signal conditioning unit in PDMG-RH GIS PD system", "GIS", "Medium"],
    # --- GIS_ARC_001 ---
    ["arc detection",       "GIS_ARC_001; GIS_PD_001", "",
     "GIS arc flash/discharge detection capability (PDMG-RH Gen3 feature)", "GIS", "High"],
    ["arc flash",           "GIS_ARC_001", "",
     "Arc flash event in GIS requiring detection and localization", "GIS", "High"],
    ["flashover",           "GIS_ARC_001", "",
     "GIS internal flashover event (arc discharge)", "GIS", "High"],
    ["GIS flashover",       "GIS_ARC_001", "",
     "Flashover in gas-insulated switchgear compartment", "GIS", "High"],
    ["arc localization",    "GIS_ARC_001", "",
     "Pinpointing arc location in GIS compartment to enable fast bay re-energization", "GIS", "High"],
    # --- BRK_HEALTH_001 ---
    ["QBCM",                "BRK_HEALTH_001", "",
     "Qualitrol QBCM circuit breaker condition monitor", "Circuit breaker", "High"],
    ["BCM",                 "BRK_HEALTH_001", "",
     "Circuit breaker condition monitor (generic / Qualitrol product)", "Circuit breaker", "High"],
    ["breaker condition monitor", "BRK_HEALTH_001", "",
     "Online condition monitor for high-voltage circuit breakers", "HV circuit breaker", "High"],
    ["CBT 200",             "BRK_HEALTH_001", "",
     "Hathaway CBT 200 circuit breaker monitor", "Circuit breaker", "Medium"],
    ["BCM 200",             "BRK_HEALTH_001", "",
     "Hathaway BCM 200/200E circuit breaker monitor", "Circuit breaker", "Medium"],
    # --- PQ_CLASSA_001 ---
    ["INFORMA PMD-A",       "PQ_CLASSA_001", "MET_PQ_CLASS",
     "Qualitrol INFORMA PMD-A Class A power quality monitor", "PCC / bus / feeder", "High"],
    ["PMD-A",               "PQ_CLASSA_001", "MET_PQ_CLASS",
     "INFORMA PMD-A power quality monitor (abbreviated)", "PCC / feeder", "High"],
    ["INFORMA",             "PQ_CLASSA_001", "",
     "Qualitrol INFORMA series power quality monitors", "PCC / bus", "Medium"],
    ["power system monitor", "DFR_DDR_001; PQ_CLASSA_001", "",
     "Multifunction substation monitor (IDM+ / INFORMA PMD-A)", "Substation", "Medium"],
    # --- TR_DGA_001 ---
    ["TM1",                 "TR_DGA_001", "MET_DGA_GAS_COUNT",
     "Qualitrol Serveron TM1 single-gas DGA monitor", "Transformer", "High"],
    ["TM3",                 "TR_DGA_001", "MET_DGA_GAS_COUNT",
     "Qualitrol Serveron TM3 3-gas DGA monitor", "Transformer", "High"],
    ["TM8",                 "TR_DGA_001", "MET_DGA_GAS_COUNT",
     "Qualitrol Serveron TM8 8-gas DGA monitor", "Transformer", "High"],
    ["Serveron",            "TR_DGA_001", "",
     "Qualitrol Serveron online DGA monitor series", "Transformer", "High"],
    ["online oil gas analysis", "TR_DGA_001", "MET_DGA_GAS_SPECIES",
     "Online dissolved gas analysis in transformer oil", "Transformer", "High"],
    # --- WAMS_001 ---
    ["inter-area oscillation", "WAMS_001; PMU_001", "",
     "Low-frequency power swing between grid areas (WAMS use case)", "Transmission grid", "High"],
    ["power swing monitoring", "WAMS_001; PMU_001", "",
     "Wide-area monitoring of power system oscillations", "Transmission grid", "High"],
    ["frequency disturbance", "WAMS_001; DFR_DDR_001", "",
     "System frequency deviation event (DFR or WAMS scope)", "Grid", "Medium"],
]

# ── Sheet 07: New Product Models ────────────────────────────────────────────
# ProductID | Model | FamilyID | Family | Scenarios | AssetType | Description |
# Standards | Protocols | DefaultQtyRule | ReqAccessories | OptAccessories | DatasheetURL |
# SourceOwner | Status | Notes
SOURCE_NOTE = "Verified via Tavily web research + Gemba Samples BOQ/product manuals."
NEW_PRODUCTS = [
    # ── PF_DFR variants ──────────────────────────────────────────────────────
    [
        "PROD_PF_DFR_03", "IDM+ 9",
        "PF_DFR", "Digital Fault / Disturbance Recorder",
        "DFR_DDR_001; PMU_001; FAULT_LOC_001; PQ_CLASSA_001",
        "Substation / feeder / plant bus",
        "IDM+ 9: 9 analog and 32 digital channels (3U, 19-inch rack). "
        "Smallest IDM+ variant. Functions: DFR (512 samples/cycle), DDR, "
        "optional PMU (IEEE C37.118 2014), optional TWS card (±60m FL), "
        "optional Class A PQ (IEC 61000-4-30 Class A Ed.2).",
        "IEEE C37.118-2014; IEC 61000-4-30 Class A; IEC 61850",
        "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus",
        "QR_DFR_001", "", "GPS antenna; IRIG-B module; GPS splitter",
        "https://www.qualitrolcorp.com/products/IDMPlus",
        SOURCE_NOTE, "Verified",
        "BOQ shorthand: IDM+ 9A/32D. 3U rackmount. Weight 15 kg. "
        "Sampling: 30.7 kHz@60Hz / 25.6 kHz@50Hz. Storage: 4 GB (opt. 16 GB). "
        "GPS time tagging: 100 ns.",
    ],
    [
        "PROD_PF_DFR_04", "IDM+ 18",
        "PF_DFR", "Digital Fault / Disturbance Recorder",
        "DFR_DDR_001; PMU_001; FAULT_LOC_001; PQ_CLASSA_001",
        "Substation / feeder / plant bus",
        "IDM+ 18: 18 analog and 64 digital channels (6U, 19-inch rack). "
        "Mid-range IDM+ variant for substations with more feeders. "
        "Same optional functions as IDM+ 9 (PMU, TWS, Class A PQ).",
        "IEEE C37.118-2014; IEC 61000-4-30 Class A; IEC 61850",
        "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus",
        "QR_DFR_001", "", "GPS antenna; IRIG-B; TWS card (optional)",
        "https://www.qualitrolcorp.com/products/IDMPlus",
        SOURCE_NOTE, "Verified",
        "BOQ shorthand: IDM+ 18A/64D. 6U rackmount. Weight 23 kg. "
        "Two 9-channel analog blocks. Suitable for 5–8 feeder substations.",
    ],
    [
        "PROD_PF_DFR_05", "IDM+ 36",
        "PF_DFR", "Digital Fault / Disturbance Recorder",
        "DFR_DDR_001; PMU_001; FAULT_LOC_001; PQ_CLASSA_001",
        "Substation / feeder / plant bus",
        "IDM+ 36: 36 analog and 128 digital channels (6U, 19-inch rack). "
        "Largest standard IDM+ variant for substations with many feeders. "
        "Same optional functions (PMU, TWS, Class A PQ).",
        "IEEE C37.118-2014; IEC 61000-4-30 Class A; IEC 61850",
        "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus",
        "QR_DFR_001", "", "GPS antenna; IRIG-B; TWS card (optional)",
        "https://www.qualitrolcorp.com/products/IDMPlus",
        SOURCE_NOTE, "Verified",
        "BOQ shorthand: IDM+ 36A/128D (or IDM+ 36A/64D for 64-DI variant). "
        "6U rackmount. Four 9-channel analog blocks. "
        "Suitable for 10–12 feeder substations.",
    ],
    # ── PF_TWS ───────────────────────────────────────────────────────────────
    [
        "PROD_PF_TWS_04", "TWS FL-8",
        "PF_TWS", "Traveling Wave Fault Locator",
        "FAULT_LOC_001",
        "High-voltage transmission line / centralized relay room",
        "TWS FL-8: modular traveling wave fault locator for centralized relay rooms. "
        "Monitors 2, 4, 6, or 8 transmission line ends (expandable in steps of 2). "
        "Uses linear coupler transducers on CT secondary wiring. "
        "GPS time tagging for double-ended fault location. "
        "Best accuracy ±60 m (±195 ft) independent of line length and impedance.",
        "IEC 61850; GPS IEEE 1588 (optional)",
        "DNP 3.0; iQ+ Master Station; serial; TCP/IP",
        "QR_TWS_001", "Linear couplers (LC); GPS antenna and cable", "iQ+ software",
        "http://hvtest.co.za/Company/PDF/Qualitrol/TWS_FL8_FL1.pdf",
        SOURCE_NOTE, "Verified",
        "3U, 19-inch rack. 132.5 mm H × 487 mm W × 362.2 mm D. 11 kg. "
        "Sample rate: 20 MHz. Channels: 3 per line (one per phase), 12-bit ADC. "
        "BOQ note: 'TWS FL-2' = FL-8 unit configured for 2 lines (not a separate model).",
    ],
    [
        "PROD_PF_TWS_05", "TWS FL-1",
        "PF_TWS", "Traveling Wave Fault Locator",
        "FAULT_LOC_001",
        "High-voltage transmission line / distributed substation",
        "TWS FL-1: fixed-format single-line traveling wave fault locator. "
        "Designed for distributed substations monitoring one line end. "
        "Same technology as FL-8; fixed 1-line capacity. "
        "Accuracy ±60 m (±195 ft).",
        "",
        "DNP 3.0; iQ+ Master Station",
        "QR_TWS_001", "Linear couplers (1 set); GPS antenna", "",
        "http://hvtest.co.za/Company/PDF/Qualitrol/TWS_FL8_FL1.pdf",
        SOURCE_NOTE, "Verified",
        "Fixed format (not modular). One line module only. "
        "Used when each substation monitors only one transmission line end. "
        "FL-8 is preferred for substations with multiple line ends.",
    ],
    # ── PF_PQ - INFORMA PMD-A ───────────────────────────────────────────────
    [
        "PROD_PF_PQ_06", "INFORMA PMD-A",
        "PF_PQ", "Power Quality Recorder",
        "PQ_CLASSA_001",
        "PCC / bus / feeder / distribution substation",
        "Qualitrol INFORMA PMD-A: Class A power quality and fault recording device. "
        "Single device eliminates 90% of analytical time. "
        "9 configurable analog channels (3U) or 9/18/27/36 channels (6U). "
        "Compliant with IEC 61000-4-30 Class A Ed.2 and IEC 61000-3-6/7 (EN 50160).",
        "IEC 61000-4-30 Class A Ed.2; IEC 61000-3-6; IEC 61000-3-7; EN 50160",
        "Serial; Ethernet; Modbus; IEC 61850 (optional)",
        "QR_PQ_001", "", "GPS time sync module",
        "http://www.hvtest.co.za/Company/PDF/Qualitrol/PMD_A.pdf",
        SOURCE_NOTE, "Verified",
        "3U (9ch) or 6U (9-36ch). Sampling: 25.6 kHz@50Hz, 30.72 kHz@60Hz. "
        "Storage: 4 GB CF (opt. 16 GB). Calibration: once every 5 years. "
        "Primary PQ-focused product (vs IDM+ which is DFR-focused).",
    ],
    # ── PF_PMU - IDM+ with PMU ────────────────────────────────────────────
    [
        "PROD_PF_PMU_04", "IDM+ (PMU option)",
        "PF_PMU", "Phasor Measurement Unit",
        "PMU_001; DFR_DDR_001; WAMS_001",
        "Transmission grid measurement point / substation bus / feeder",
        "IDM+ with PMU option enabled: provides synchrophasor streaming per "
        "IEEE C37.118-2014 alongside full DFR/DDR functions. "
        "Available in 9, 18, or 36 analog channel variants. "
        "Used in FMS+WAMS deployments across GCC and international utilities.",
        "IEEE C37.118-2014; IEC 61850; IEC 61000-4-30 Class A",
        "IEC 61850 GOOSE; DNP 3.0; Modbus TCP; C37.118 streaming",
        "QR_PMU_001", "GPS antenna; iQ+ Master Station software", "",
        "https://www.qualitrolcorp.com/products/IDMPlus",
        SOURCE_NOTE, "Verified",
        "PMU is an optional add-on license/card for IDM+ 9/18/36. "
        "Reporting rates: 25 or 50 fps at 50 Hz; 30 or 60 fps at 60 Hz (typical). "
        "Used by GCCIA, OETC, Scottish Power, Statnett, EGAT, etc.",
    ],
]

# ── Sheet 08: New Product Parameters ────────────────────────────────────────
# Product ID | Parameter Name | Value/Range | Unit | Data Type | Notes
NEW_PARAMS = [
    # ── IDM+ 9 ──
    ["PROD_PF_DFR_03", "Analog Channels", "9", "count", "integer", "3U, single analog card block"],
    ["PROD_PF_DFR_03", "Digital Channels", "32", "count", "integer", "32 digital inputs standard"],
    ["PROD_PF_DFR_03", "Form Factor", "3U (133 mm H × 483 mm W)", "text", "text", "19-inch rack mountable"],
    ["PROD_PF_DFR_03", "DFR Sampling Rate", "512 samples/cycle (30.7 kHz@60Hz; 25.6 kHz@50Hz)", "Hz", "text", "Selectable: 512/256/128/64/32 s/c"],
    ["PROD_PF_DFR_03", "Analog Resolution", "20-bit current; 16-bit voltage/DC", "bits", "text", ""],
    ["PROD_PF_DFR_03", "Accuracy", "±0.1% full scale", "%", "text", ""],
    ["PROD_PF_DFR_03", "Data Storage", "4 GB CompactFlash standard; opt 8/16 GB", "GB", "text", ""],
    ["PROD_PF_DFR_03", "Time Synchronization", "GPS (100 ns); IRIG-B/J; 1 pps", "ns", "text", ""],
    ["PROD_PF_DFR_03", "Operating Temperature", "-5 to +50°C operating; -30 to +70°C storage", "°C", "text", ""],
    ["PROD_PF_DFR_03", "Communication Protocol", "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus", "text", "text", ""],
    ["PROD_PF_DFR_03", "Optional Functions", "PMU (IEEE C37.118-2014); TWS card (±60m FL); Class A PQ", "text", "text", "Each option ordered separately"],
    ["PROD_PF_DFR_03", "Asset Type", "Substation bus / feeder / plant bus", "text", "text", ""],
    # ── IDM+ 18 ──
    ["PROD_PF_DFR_04", "Analog Channels", "18", "count", "integer", "6U, two 9-channel analog blocks"],
    ["PROD_PF_DFR_04", "Digital Channels", "64", "count", "integer", "64 digital inputs standard"],
    ["PROD_PF_DFR_04", "Form Factor", "6U (267 mm H × 483 mm W)", "text", "text", "19-inch rack mountable"],
    ["PROD_PF_DFR_04", "DFR Sampling Rate", "512 samples/cycle (30.7 kHz@60Hz; 25.6 kHz@50Hz)", "Hz", "text", ""],
    ["PROD_PF_DFR_04", "Data Storage", "4 GB CompactFlash standard; opt 8/16 GB", "GB", "text", ""],
    ["PROD_PF_DFR_04", "Time Synchronization", "GPS (100 ns); IRIG-B/J; 1 pps", "ns", "text", ""],
    ["PROD_PF_DFR_04", "Communication Protocol", "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus", "text", "text", ""],
    ["PROD_PF_DFR_04", "Optional Functions", "PMU (IEEE C37.118-2014); TWS card (±60m FL); Class A PQ", "text", "text", ""],
    ["PROD_PF_DFR_04", "Operating Temperature", "-5 to +50°C operating", "°C", "text", ""],
    ["PROD_PF_DFR_04", "Asset Type", "Substation bus / feeder / plant bus", "text", "text", ""],
    # ── IDM+ 36 ──
    ["PROD_PF_DFR_05", "Analog Channels", "36", "count", "integer", "6U, four 9-channel analog blocks"],
    ["PROD_PF_DFR_05", "Digital Channels", "128", "count", "integer", "128 digital inputs standard"],
    ["PROD_PF_DFR_05", "Form Factor", "6U (267 mm H × 483 mm W)", "text", "text", "19-inch rack mountable"],
    ["PROD_PF_DFR_05", "DFR Sampling Rate", "512 samples/cycle (30.7 kHz@60Hz; 25.6 kHz@50Hz)", "Hz", "text", ""],
    ["PROD_PF_DFR_05", "Data Storage", "4 GB standard; opt 8/16 GB", "GB", "text", ""],
    ["PROD_PF_DFR_05", "Time Synchronization", "GPS (100 ns); IRIG-B/J; 1 pps", "ns", "text", ""],
    ["PROD_PF_DFR_05", "Communication Protocol", "IEC 61850; IEC 60870-5-104; DNP 3.0; Modbus", "text", "text", ""],
    ["PROD_PF_DFR_05", "Optional Functions", "PMU (IEEE C37.118-2014); TWS card (±60m FL); Class A PQ", "text", "text", ""],
    ["PROD_PF_DFR_05", "Operating Temperature", "-5 to +50°C operating", "°C", "text", ""],
    ["PROD_PF_DFR_05", "Asset Type", "Substation bus / feeder / plant bus", "text", "text", ""],
    # ── TWS FL-8 ──
    ["PROD_PF_TWS_04", "Lines Monitored per Device", "2, 4, 6, or 8 (start 2, expand in steps of 2)", "count", "text", "FL-8 starts with 2-line module"],
    ["PROD_PF_TWS_04", "Fault Location Accuracy", "±60 m (±195 ft), independent of line length/impedance", "m", "text", "Best accuracy up to 1000 km lines"],
    ["PROD_PF_TWS_04", "Sample Rate", "20 MHz (also 10/5/2.5/1.25 MHz selectable)", "MHz", "text", ""],
    ["PROD_PF_TWS_04", "Channels per Line", "3 (one per phase) from external linear coupler", "count", "integer", "12-bit ADC"],
    ["PROD_PF_TWS_04", "Time Synchronization", "GPS", "text", "text", "Required for double-ended fault location"],
    ["PROD_PF_TWS_04", "Communication Protocol", "DNP 3.0; iQ+ Master Station; TCP/IP", "text", "text", ""],
    ["PROD_PF_TWS_04", "Form Factor", "3U (132.5 mm H × 487 mm W × 362.2 mm D); 11 kg", "text", "text", ""],
    ["PROD_PF_TWS_04", "Asset Type", "HV transmission line (line end at substation)", "text", "text", ""],
    ["PROD_PF_TWS_04", "Installation Method", "Linear couplers on CT secondary; no line outage", "text", "text", ""],
    # ── TWS FL-1 ──
    ["PROD_PF_TWS_05", "Lines Monitored per Device", "1", "count", "integer", "Fixed format, single-line only"],
    ["PROD_PF_TWS_05", "Fault Location Accuracy", "±60 m (±195 ft)", "m", "text", ""],
    ["PROD_PF_TWS_05", "Time Synchronization", "GPS", "text", "text", ""],
    ["PROD_PF_TWS_05", "Communication Protocol", "DNP 3.0; iQ+ Master Station", "text", "text", ""],
    ["PROD_PF_TWS_05", "Application", "Distributed substations monitoring one line end", "text", "text", ""],
    ["PROD_PF_TWS_05", "Asset Type", "HV transmission line (single end)", "text", "text", ""],
    # ── INFORMA PMD-A ──
    ["PROD_PF_PQ_06", "Power Quality Class", "IEC 61000-4-30 Class A Edition 2.0", "text", "text", "Third-party certified"],
    ["PROD_PF_PQ_06", "Analog Channels", "9 (3U) or 9/18/27/36 (6U, configurable)", "count", "text", ""],
    ["PROD_PF_PQ_06", "Sampling Rate", "25.6 kHz@50Hz; 30.72 kHz@60Hz", "Hz", "text", ""],
    ["PROD_PF_PQ_06", "Bandwidth", "25 to 4600 Hz ±0.5 dB", "Hz", "text", ""],
    ["PROD_PF_PQ_06", "Standards", "IEC 61000-4-30 Class A; IEC 61000-3-6/7; EN 50160", "text", "text", ""],
    ["PROD_PF_PQ_06", "Data Storage", "4 GB CF standard; opt 8/16 GB", "GB", "text", ""],
    ["PROD_PF_PQ_06", "Communication Protocol", "Serial; Ethernet; Modbus; IEC 61850 (opt.)", "text", "text", ""],
    ["PROD_PF_PQ_06", "Asset Type", "PCC / bus / feeder / distribution substation", "text", "text", ""],
    # ── IDM+ (PMU option) ──
    ["PROD_PF_PMU_04", "Analog channels (max)", "9 / 18 / 36 (per IDM+ variant)", "count", "text", ""],
    ["PROD_PF_PMU_04", "PMU Standard", "IEEE C37.118-2014", "text", "text", ""],
    ["PROD_PF_PMU_04", "PMU Reporting Rate", "25 or 50 fps at 50 Hz; 30 or 60 fps at 60 Hz (typical)", "fps", "text", ""],
    ["PROD_PF_PMU_04", "Time Synchronization", "GPS (100 ns)", "ns", "text", ""],
    ["PROD_PF_PMU_04", "Communication Protocol", "IEEE C37.118 streaming; IEC 61850 GOOSE; DNP 3.0", "text", "text", ""],
    ["PROD_PF_PMU_04", "Optional Add-ons", "TWS card (fault location); Class A PQ", "text", "text", ""],
    ["PROD_PF_PMU_04", "Asset Type", "Transmission grid measurement point / substation bus", "text", "text", ""],
]

# ── WRITE TO EXCEL ─────────────────────────────────────────────────────────
def write_all(wb, out_path):
    counts = {
        "s03_added": 0, "s04_added": 0, "s05_added": 0,
        "s07_added": 0, "s08_added": 0,
    }

    # ── Sheet 03 ──
    ws3 = wb["03_Scenario_Master"]
    hdr_row, hdr_map = get_header_row(ws3)
    existing_ids = existing_values(ws3, 0)
    for scen in NEW_SCENARIOS:
        if scen[0].lower() not in existing_ids:
            ws3.append(scen)
            print(f"  [S03 ADD] {scen[0]}")
            counts["s03_added"] += 1
        else:
            print(f"  [S03 SKIP] {scen[0]} (exists)")

    # ── Sheet 04 ──
    ws4 = wb["04_Metric_Dictionary"]
    existing_met = existing_values(ws4, 0)
    for met in NEW_METRICS:
        if met[0].lower() not in existing_met:
            ws4.append(met)
            print(f"  [S04 ADD] {met[0]}")
            counts["s04_added"] += 1
        else:
            print(f"  [S04 SKIP] {met[0]} (exists)")

    # ── Sheet 05 ──
    ws5 = wb["05_Synonym_Mapping"]
    existing_syn = existing_values(ws5, 0)
    for syn in NEW_SYNONYMS:
        if syn[0].lower() not in existing_syn:
            ws5.append(syn)
            print(f"  [S05 ADD] {syn[0]}")
            counts["s05_added"] += 1
        else:
            print(f"  [S05 SKIP] {syn[0]} (exists)")

    # ── Sheet 07 ──
    ws7 = wb["07_Product_Master_Template"]
    existing_prod = existing_values(ws7, 0)
    for prod in NEW_PRODUCTS:
        if prod[0].lower() not in existing_prod:
            ws7.append(prod)
            print(f"  [S07 ADD] {prod[0]} – {prod[1]}")
            counts["s07_added"] += 1
        else:
            print(f"  [S07 SKIP] {prod[0]} (exists)")

    # ── Sheet 08 ──
    ws8 = wb["08_Product_Parameter_Template"]
    # Build existing set as (prod_id, param_name)
    existing_param = set()
    for row in ws8.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            existing_param.add((str(row[0]).strip().lower(), str(row[1]).strip().lower()))
    for param in NEW_PARAMS:
        key = (param[0].lower(), param[1].lower())
        if key not in existing_param:
            ws8.append(param)
            counts["s08_added"] += 1
        else:
            print(f"  [S08 SKIP] {param[0]} / {param[1]} (exists)")

    wb.save(out_path)
    return counts


# ── MAIN ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not MASTER.exists():
        print(f"ERROR: Master file not found: {MASTER}")
        sys.exit(1)

    print(f"Backing up master → {BACKUP.name} ...")
    shutil.copy2(MASTER, BACKUP)

    print(f"Loading workbook ...")
    wb, tmp = load_wb()

    print("\nWriting new content ...")
    counts = write_all(wb, tmp)

    # Copy temp back to master
    import shutil as _sh
    _sh.copy2(tmp, MASTER)
    try:
        tmp.unlink(missing_ok=True)
    except Exception:
        pass

    print(f"""
Done.
  Sheet 03 (Scenarios) added : {counts['s03_added']}
  Sheet 04 (Metrics)   added : {counts['s04_added']}
  Sheet 05 (Synonyms)  added : {counts['s05_added']}
  Sheet 07 (Products)  added : {counts['s07_added']}
  Sheet 08 (Params)    added : {counts['s08_added']}

Saved to: {MASTER}
Backup  : {BACKUP}
""")
