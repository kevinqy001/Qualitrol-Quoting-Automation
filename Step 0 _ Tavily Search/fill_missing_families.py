"""Fill missing product families in the master data package via Step 0 Tavily research.

Targets only the families that have no product models yet:
    PF_DFR       - Digital Fault / Disturbance Recorder  (family exists, no models)
    PF_TR_TEMP   - Transformer Temperature Monitor       (family exists, no models)
    PF_BUSHING   - Bushing Monitor                       (family exists, no models)
    PF_AUX_SENSOR- Transformer Auxiliary Sensors         (family exists, no models)
    PF_TWS       - Traveling Wave Fault Locator          (completely new family)

Safety guarantees:
  * Existing products in sheets 07 / 08 are PRESERVED; new rows are only APPENDED.
  * Sheet 06 gets PF_TWS appended (existing rows untouched).
  * A timestamped backup is created before any write.
  * If the master file is locked (open in Excel), a _patched copy is saved instead.

Usage (from repo root):
    python "Step 0 _ Tavily Search/fill_missing_families.py"
    python "Step 0 _ Tavily Search/fill_missing_families.py" --dry-run
    python "Step 0 _ Tavily Search/fill_missing_families.py" --plan-only
"""

from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402

from qualitrol_core import config, io_utils  # noqa: E402

# ── Constants ──────────────────────────────────────────────────────────────── #

MISSING_FAMILIES = ["PF_DFR", "PF_TR_TEMP", "PF_BUSHING", "PF_AUX_SENSOR", "PF_TWS"]

# New PF_TWS row to inject into sheet 06 (all 8 columns in header order)
PF_TWS_ROW = (
    "PF_TWS",
    "Traveling Wave Fault Locator",
    "FAULT_LOC_001",
    "Transmission line / feeder",
    (
        "Traveling wave detection; single-end and double-end fault location; "
        "GPS/IEEE 1588 time sync; line protection relay integration; "
        "sub-cycle fault location accuracy"
    ),
    "QR_FAULT_LOC_001",
    (
        "Line count; line length (km); wave propagation velocity; "
        "GPS time-sync availability; CT/VT channel count; "
        "relay integration requirement"
    ),
    (
        "Qualitrol TWS FL-2 product confirmed in reference quotes. "
        "Confirm whether integrated with DFR or standalone. "
        "Verify channel count and line configuration."
    ),
)

BACKUP_DIR = config.OUTPUT_DIR / "_product_catalog" / "backups"
CATALOG_JSON = config.OUTPUT_DIR / "_product_catalog" / "step0_product_catalog.json"


# ── Helpers ────────────────────────────────────────────────────────────────── #

def _find_header_row(ws, first_col: str) -> int:
    target = first_col.strip().lower()
    for row in ws.iter_rows():
        if row[0].value is not None and str(row[0].value).strip().lower() == target:
            return row[0].row
    raise ValueError(f"Header '{first_col}' not found in sheet '{ws.title}'")


def _header_map(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(row=header_row, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if str(ws.cell(row=header_row, column=c).value or "").strip()
    }


def _existing_family_ids(ws, header_row: int) -> set[str]:
    col = _header_map(ws, header_row).get("Product Family ID", 1)
    return {
        str(ws.cell(row=r, column=col).value or "").strip()
        for r in range(header_row + 1, ws.max_row + 1)
        if ws.cell(row=r, column=col).value
    }


def _existing_product_ids(ws, header_row: int) -> set[str]:
    col = _header_map(ws, header_row).get("Product ID", 1)
    return {
        str(ws.cell(row=r, column=col).value or "").strip()
        for r in range(header_row + 1, ws.max_row + 1)
        if ws.cell(row=r, column=col).value
    }


def _write_row(ws, row_idx: int, hmap: dict[str, int], values: dict) -> None:
    for header, value in values.items():
        col = hmap.get(header)
        if col:
            ws.cell(row=row_idx, column=col, value=value)


def _existing_match_defaults(ws, header_row: int) -> dict[str, tuple]:
    hmap = _header_map(ws, header_row)
    pid_col = hmap.get("Parameter ID")
    mt_col = hmap.get("Match Type")
    mp_col = hmap.get("Match Priority")
    defaults: dict[str, tuple] = {}
    if not (pid_col and mt_col and mp_col):
        return defaults
    for r in range(header_row + 1, ws.max_row + 1):
        pid = str(ws.cell(row=r, column=pid_col).value or "").strip()
        if pid and pid not in defaults:
            defaults[pid] = (
                str(ws.cell(row=r, column=mt_col).value or "").strip(),
                str(ws.cell(row=r, column=mp_col).value or "").strip(),
            )
    return defaults


# ── Step 1: prepare a temp data package with PF_TWS injected ──────────────── #

def _make_temp_package_with_tws(master_path: Path) -> Path:
    """Copy master to a temp file and append the PF_TWS row if absent."""
    tmp = Path(tempfile.mktemp(suffix=".xlsx", prefix="qboq_tws_"))
    shutil.copy2(master_path, tmp)

    wb = openpyxl.load_workbook(str(tmp))
    ws6 = wb["06_Product_Family_Master"]
    h_row = _find_header_row(ws6, "Product Family ID")
    existing = _existing_family_ids(ws6, h_row)

    if "PF_TWS" not in existing:
        ws6.append(list(PF_TWS_ROW))
        print("[06] Appended PF_TWS to temp copy of sheet 06.")
    else:
        print("[06] PF_TWS already present in sheet 06 - skipping inject.")

    wb.save(str(tmp))
    wb.close()
    return tmp


# ── Step 2: run Step 0 pipeline against the temp package ─────────────────── #

def _run_step0(temp_package: Path, dry_run: bool) -> dict | None:
    """Run the Tavily pipeline for the missing families only.

    We call the internal research helpers directly (rather than pipeline.run)
    so that we can pass the temp data-package (which already has PF_TWS in it)
    instead of letting pipeline.run re-load the default master file.
    """
    step0_dir = Path(__file__).resolve().parent
    if str(step0_dir) not in sys.path:
        sys.path.insert(0, str(step0_dir))

    import catalog_excel  # sibling module

    from qualitrol_core import tavily_client, llm, io_utils
    from qualitrol_core.data_package import load_data_package
    from qualitrol_core import product_research as pr

    # Load from the TEMP copy so PF_TWS is visible to the research loop.
    dp = load_data_package(str(temp_package))
    tv = tavily_client.get_client()
    llm_client = llm.get_client()
    primary = config.SETTINGS.tavily_primary_domain

    print(f"\nTavily available : {tv.available}")
    print(f"LLM    available : {llm_client.available}")

    families = [f for f in dp.families.values() if f.family_id in set(MISSING_FAMILIES)]

    if dry_run:
        plan = pr.build_full_query_plan(dp, primary)
        print("\n[dry-run] Query plan for missing families:")
        for q in plan["family_queries"]:
            if q.get("family_id") in MISSING_FAMILIES:
                doms = ", ".join(q.get("include_domains") or ["(open web)"])
                print(f"  {q['family_id']:<16}  {q['query']}  :: {doms}")
        return None

    products: list[dict] = []
    parameters: list[dict] = []
    per_family: list[dict] = []

    for family in families:
        print(f"  Researching {family.family_id} ({family.family_name}) ...")
        queries = pr.build_family_queries(family, primary)
        merged: dict = {"results": []}
        for q in queries:
            resp = tv.search(q["query"], include_domains=q.get("include_domains") or None)
            merged["results"].extend(resp.get("results", []) or [])

        urls = pr.collect_urls(merged, primary,
                               limit=config.SETTINGS.tavily_max_urls_per_family)
        extract_resp = tv.extract(urls) if urls else {"results": []}
        context = pr._gather_text(merged, extract_resp)
        structured = pr.structure_family(llm_client, family, dp, context)

        n = len(structured["products"]) if structured else 0
        per_family.append({"family_id": family.family_id,
                            "family_name": family.family_name,
                            "urls_used": urls, "models_found": n})
        if structured:
            products.extend(structured["products"])
            parameters.extend(structured["parameters"])
        print(f"    -> {n} product(s) found")

    families_out = [
        {"family_id": f.family_id, "family_name": f.family_name,
         "applicable_scenarios": f.applicable_scenarios,
         "primary_asset_type": f.primary_asset_type,
         "typical_capabilities": f.typical_capabilities,
         "default_quantity_rule_id": f.default_quantity_rule_id,
         "dependencies": f.dependencies, "notes": f.notes}
        for f in dp.families.values()
    ]

    result = {
        "step": "0_tavily_search",
        "tavily": {"available": tv.available, "executed": True,
                   "primary_domain": primary,
                   "search_depth": config.SETTINGS.tavily_search_depth},
        "llm": {"available": llm_client.available,
                "model": config.SETTINGS.llm_deployment if llm_client.available else None},
        "query_plan": pr.build_full_query_plan(dp, primary),
        "summary": {"families": len(families_out),
                    "families_researched": len(per_family),
                    "products_found": len(products),
                    "parameters_found": len(parameters)},
        "per_family": per_family,
        "product_families": families_out,
        "products": products,
        "product_parameters": parameters,
    }

    out_path = io_utils.write_json(
        config.OUTPUT_DIR / "_product_catalog" / "step0_product_catalog.json", result
    )
    result["_output_path"] = str(out_path)
    return result


# ── Step 3: append-merge into master ────────────────────────────────────────── #

def _append_merge(catalog: dict, master_path: Path) -> dict:
    """Append newly discovered products/params to the master without clearing."""
    products = catalog.get("products", [])
    parameters = catalog.get("product_parameters", [])

    if not products:
        print("\n[WARN] No products found in Step 0 output - nothing to write.")
        return {"products_added": 0, "params_added": 0}

    # ── Backup ──────────────────────────────────────────────────────────── #
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"Qualitrol_BOQ_Matching_Data_Package__{stamp}.xlsx"
    shutil.copy2(master_path, backup)
    print(f"\nBackup saved: {backup.name}")

    # ── Load master (or fall back to backup copy if locked) ─────────────── #
    output_path = master_path
    try:
        wb = openpyxl.load_workbook(str(master_path))
    except PermissionError:
        print(
            "\n[WARNING] Master file is open in Excel.\n"
            "          Writing to a _patched copy instead.\n"
            "          Close Excel then replace master with the patched file."
        )
        wb = openpyxl.load_workbook(str(backup))
        output_path = BACKUP_DIR / f"Qualitrol_BOQ_Matching_Data_Package__patched_{stamp}.xlsx"

    ws6 = wb["06_Product_Family_Master"]
    ws_prod = wb["07_Product_Master_Template"]
    ws_param = wb["08_Product_Parameter_Template"]

    h6 = _find_header_row(ws6, "Product Family ID")
    h7 = _find_header_row(ws_prod, "Product ID")
    h8 = _find_header_row(ws_param, "Product ID")

    existing_families = _existing_family_ids(ws6, h6)
    existing_products = _existing_product_ids(ws_prod, h7)
    match_defaults = _existing_match_defaults(ws_param, h8)

    prod_hmap = _header_map(ws_prod, h7)
    param_hmap = _header_map(ws_param, h8)

    # ── Sheet 06: append PF_TWS if still missing ────────────────────────── #
    if "PF_TWS" not in existing_families:
        ws6.append(list(PF_TWS_ROW))
        print("[06] Appended PF_TWS to sheet 06.")
    else:
        print("[06] PF_TWS already in sheet 06.")

    # ── Sheet 07: append only truly new product IDs ──────────────────────── #
    products_added = 0
    r07 = ws_prod.max_row + 1
    for p in products:
        pid = p.get("product_id", "")
        if pid in existing_products:
            print(f"[07] SKIP (already exists): {pid} - {p.get('model', '')}")
            continue
        _write_row(ws_prod, r07, prod_hmap, {
            "Product ID": pid,
            "Product Model": p.get("model", ""),
            "Product Family ID": p.get("family_id", ""),
            "Product Family": p.get("family_name", ""),
            "Applicable Scenario IDs": "; ".join(p.get("applicable_scenarios", [])),
            "Primary Asset Type": p.get("primary_asset_type", ""),
            "Product Description": p.get("description", ""),
            "Supported Standards": p.get("supported_standards", ""),
            "Communication Protocols": p.get("protocols", ""),
            "Default Quantity Rule ID": p.get("default_quantity_rule_id", ""),
            "Datasheet URL": p.get("datasheet_url", ""),
            "Source Owner": "Tavily research (pending review)",
            "Status": p.get("status", "Candidate"),
            "Notes": p.get("notes", ""),
        })
        print(f"[07] Added: {pid} - {p.get('model', '')} ({p.get('family_id', '')})")
        r07 += 1
        products_added += 1

    # ── Sheet 08: append parameters for new products ─────────────────────── #
    params_added = 0
    new_pids = {p.get("product_id") for p in products if p.get("product_id") not in existing_products}
    r08 = ws_param.max_row + 1
    for prm in parameters:
        if prm.get("product_id") not in new_pids:
            continue
        mid = prm.get("metric_id", "")
        mt, mp = match_defaults.get(mid, ("", ""))
        _write_row(ws_param, r08, param_hmap, {
            "Product ID": prm["product_id"],
            "Product Model": prm.get("model", ""),
            "Product Family ID": prm.get("family_id", ""),
            "Parameter ID": mid,
            "Parameter Name": prm.get("parameter_name", ""),
            "Min Value": prm.get("min_value"),
            "Max Value": prm.get("max_value"),
            "Supported Value / Text": prm.get("supported_value", ""),
            "Unit": prm.get("unit", ""),
            "Match Type": mt,
            "Match Priority": mp,
            "Evidence Source / Datasheet URL": prm.get("datasheet_url", ""),
            "Notes": prm.get("notes", ""),
        })
        r08 += 1
        params_added += 1

    wb.save(str(output_path))
    wb.close()

    result = {
        "output_path": str(output_path),
        "backup": str(backup),
        "products_added": products_added,
        "params_added": params_added,
    }

    if output_path != master_path:
        print(f"\nPatched file: {output_path.name}")
        print("To apply:")
        print("  1. Close Qualitrol_BOQ_Matching_Data_Package.xlsx in Excel")
        print(f"  2. Replace it with: {output_path}")
        print("  3. Restart web server: python app.py")
    else:
        print(f"\nMaster updated: {master_path.name}")
        print("Restart the web server to reload: python app.py")

    return result


# ── Main ───────────────────────────────────────────────────────────────────── #

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Fill missing Qualitrol product families via Tavily Step 0"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Show query plan; do not run searches or write files.")
    parser.add_argument("--plan-only", action="store_true",
                        help="Alias for --dry-run.")
    args = parser.parse_args(argv)
    dry = args.dry_run or args.plan_only

    master = config.DATA_PACKAGE_PATH
    print("=" * 65)
    print("Fill Missing Product Families - Step 0 Append-Merge")
    print("=" * 65)
    print(f"Target families : {', '.join(MISSING_FAMILIES)}")
    print(f"Master file     : {master.name}")
    print(f"Mode            : {'DRY RUN' if dry else 'LIVE'}")

    # Step 1: temp copy with PF_TWS injected (only needed for research)
    if not dry:
        print("\n[Step 1] Preparing temp data package with PF_TWS ...")
        tmp = _make_temp_package_with_tws(master)
    else:
        tmp = master  # point directly at master for dry-run plan

    # Step 2: run Tavily
    print("\n[Step 2] Running Step 0 Tavily research ...")
    catalog = _run_step0(tmp, dry_run=dry)

    if dry:
        print("\n[dry-run] Skipping write step.")
        return 0

    # Clean up temp file
    try:
        tmp.unlink(missing_ok=True)
    except Exception:
        pass

    if catalog is None:
        print("[ERROR] Step 0 returned no result.")
        return 1

    tv = catalog.get("tavily", {})
    s = catalog.get("summary", {})
    print(f"\nTavily executed={tv.get('executed')} | "
          f"products found={s.get('products_found', 0)} | "
          f"params={s.get('parameters_found', 0)}")

    if not catalog.get("products"):
        print("[WARN] No products discovered. Check Tavily API key and connectivity.")
        return 0

    # Step 3: append-merge
    print("\n[Step 3] Appending new products to master data package ...")
    result = _append_merge(catalog, master)

    print(f"\n{'='*65}")
    print(f"Products added : {result['products_added']}")
    print(f"Parameters added: {result['params_added']}")
    print(f"Backup         : {Path(result['backup']).name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
