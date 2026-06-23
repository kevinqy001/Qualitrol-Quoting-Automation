"""Write the researched catalog into a standalone candidate workbook.

Produces a NEW .xlsx (never the master data package) with three sheets that
mirror 06/07/08 so a human can review and paste verified rows back into
``Qualitrol_BOQ_Matching_Data_Package.xlsx``.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_FAMILY_COLS = [
    "Product Family ID", "Product Family", "Applicable Scenario IDs",
    "Primary Asset Type", "Typical Capabilities", "Default Quantity Rule ID",
    "Dependencies / Required Inputs", "Commercial / Engineering Notes",
]
_PRODUCT_COLS = [
    "Product ID", "Product Model", "Product Family ID", "Product Family",
    "Applicable Scenario IDs", "Primary Asset Type", "Product Description",
    "Supported Standards", "Communication Protocols", "Default Quantity Rule ID",
    "Datasheet URL", "Status", "Notes",
]
_PARAM_COLS = [
    "Product ID", "Product Model", "Product Family ID", "Parameter ID",
    "Parameter Name", "Min Value", "Max Value", "Supported Value / Text",
    "Unit", "Match Type", "Match Priority", "Evidence Source / Datasheet URL",
    "Notes",
]


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _write_header(ws, cols: list[str]) -> None:
    ws.append(cols)
    _style_header(ws, len(cols))
    ws.freeze_panes = "A2"


def write_catalog_workbook(path: str | Path, families: list[dict],
                           products: list[dict], parameters: list[dict],
                           query_plan: dict) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws_fam = wb.active
    ws_fam.title = "06_Product_Family_Master"
    _write_header(ws_fam, _FAMILY_COLS)
    for f in families:
        ws_fam.append([
            f.get("family_id", ""), f.get("family_name", ""),
            "; ".join(f.get("applicable_scenarios", [])),
            f.get("primary_asset_type", ""), f.get("typical_capabilities", ""),
            f.get("default_quantity_rule_id", ""), f.get("dependencies", ""),
            f.get("notes", ""),
        ])

    ws_prod = wb.create_sheet("07_Product_Master_Template")
    _write_header(ws_prod, _PRODUCT_COLS)
    for p in products:
        ws_prod.append([
            p.get("product_id", ""), p.get("model", ""), p.get("family_id", ""),
            p.get("family_name", ""), "; ".join(p.get("applicable_scenarios", [])),
            p.get("primary_asset_type", ""), p.get("description", ""),
            p.get("supported_standards", ""), p.get("protocols", ""),
            p.get("default_quantity_rule_id", ""), p.get("datasheet_url", ""),
            p.get("status", "Candidate"), p.get("notes", ""),
        ])

    ws_param = wb.create_sheet("08_Product_Parameter_Template")
    _write_header(ws_param, _PARAM_COLS)
    for prm in parameters:
        ws_param.append([
            prm.get("product_id", ""), prm.get("model", ""), prm.get("family_id", ""),
            prm.get("metric_id", ""), prm.get("parameter_name", ""),
            prm.get("min_value"), prm.get("max_value"),
            prm.get("supported_value", ""), prm.get("unit", ""),
            prm.get("match_type", ""), prm.get("match_priority", ""),
            prm.get("notes", ""),
        ])

    ws_plan = wb.create_sheet("00_Query_Plan")
    _write_header(ws_plan, ["Purpose", "Query", "Include Domains", "Family ID"])
    for q in query_plan.get("discovery_queries", []):
        ws_plan.append([q.get("purpose", ""), q.get("query", ""),
                        "; ".join(q.get("include_domains", [])), ""])
    for q in query_plan.get("family_queries", []):
        ws_plan.append([q.get("purpose", ""), q.get("query", ""),
                        "; ".join(q.get("include_domains", [])),
                        q.get("family_id", "")])

    for ws, cols in (
        (ws_fam, _FAMILY_COLS), (ws_prod, _PRODUCT_COLS),
        (ws_param, _PARAM_COLS),
    ):
        for idx in range(1, len(cols) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 24

    wb.save(path)
    return path
