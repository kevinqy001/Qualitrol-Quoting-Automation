"""Backfill the master data package with Step 0's researched catalog.

Writes the discovered product models into sheet 07 and their parameters into
sheet 08 of ``Qualitrol_BOQ_Matching_Data_Package.xlsx``. Sheet 06 (families) is
left untouched (already curated). A timestamped backup is always taken first.

Safeguards:
  * Backup copy saved under outputs/_product_catalog/backups/ before any write.
  * Only the data rows below each header are replaced; title/description/header
    rows and every other sheet are preserved.
  * Match Type / Match Priority defaults from the original sheet 08 are carried
    over per Metric ID so matching semantics are not lost.
  * Writes by header NAME (not column index) so it is robust to layout changes.

Usage (from repo root):
    python "Step 0 _ Tavily Search/backfill.py"            # uses latest Step 0 JSON
    python "Step 0 _ Tavily Search/backfill.py" --dry-run  # report only, no write
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402

from qualitrol_core import config, io_utils  # noqa: E402

CATALOG_JSON = config.OUTPUT_DIR / "_product_catalog" / "step0_product_catalog.json"


def _find_header_row(ws, first_col: str = "Product ID") -> int:
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == first_col:
            return r
    raise ValueError(f"Header row '{first_col}' not found in {ws.title}")


def _header_map(ws, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(row=header_row, column=c).value or "").strip()
        if name:
            out[name] = c
    return out


def _clear_data_rows(ws, header_row: int) -> None:
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def _write_row(ws, row_idx: int, hmap: dict[str, int], values: dict) -> None:
    for header, value in values.items():
        col = hmap.get(header)
        if col:
            ws.cell(row=row_idx, column=col, value=value)


def _existing_match_defaults(ws, header_row: int) -> dict[str, tuple]:
    """Map Metric/Parameter ID -> (Match Type, Match Priority) from current 08."""
    hmap = _header_map(ws, header_row)
    pid_col = hmap.get("Parameter ID")
    mt_col = hmap.get("Match Type")
    mp_col = hmap.get("Match Priority")
    defaults: dict[str, tuple] = {}
    if not (pid_col and mt_col and mp_col):
        return defaults
    for r in range(header_row + 1, ws.max_row + 1):
        pid = str(ws.cell(row=r, column=pid_col).value or "").strip()
        if pid and pid not in defaults:
            defaults[pid] = (
                str(ws.cell(row=r, column=mt_col).value or "").strip(),
                str(ws.cell(row=r, column=mp_col).value or "").strip(),
            )
    return defaults


def backfill(catalog_path: Path = CATALOG_JSON, master_path: Path | None = None,
             dry_run: bool = False) -> dict:
    master_path = master_path or config.DATA_PACKAGE_PATH
    catalog = io_utils.read_json(catalog_path)
    products = catalog.get("products", [])
    parameters = catalog.get("product_parameters", [])
    if not products:
        raise SystemExit("No products in catalog JSON; run Step 0 first.")

    url_by_pid = {p["product_id"]: p.get("datasheet_url", "") for p in products}

    wb = openpyxl.load_workbook(master_path)
    ws_prod = wb["07_Product_Master_Template"]
    ws_param = wb["08_Product_Parameter_Template"]

    prod_hrow = _find_header_row(ws_prod)
    param_hrow = _find_header_row(ws_param)
    match_defaults = _existing_match_defaults(ws_param, param_hrow)

    summary = {
        "master": str(master_path),
        "products_written": len(products),
        "parameters_written": len(parameters),
        "match_defaults_reused": 0,
        "backup": "",
        "dry_run": dry_run,
    }

    if dry_run:
        return summary

    # 1. Backup.
    backup_dir = config.OUTPUT_DIR / "_product_catalog" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{master_path.stem}__{stamp}{master_path.suffix}"
    shutil.copy2(master_path, backup_path)
    summary["backup"] = str(backup_path)

    # 2. Sheet 07 - products.
    prod_hmap = _header_map(ws_prod, prod_hrow)
    _clear_data_rows(ws_prod, prod_hrow)
    r = prod_hrow + 1
    for p in products:
        _write_row(ws_prod, r, prod_hmap, {
            "Product ID": p["product_id"],
            "Product Model": p["model"],
            "Product Family ID": p["family_id"],
            "Product Family": p["family_name"],
            "Applicable Scenario IDs": "; ".join(p.get("applicable_scenarios", [])),
            "Primary Asset Type": p.get("primary_asset_type", ""),
            "Product Description": p.get("description", ""),
            "Supported Standards": p.get("supported_standards", ""),
            "Communication Protocols": p.get("protocols", ""),
            "Default Quantity Rule ID": p.get("default_quantity_rule_id", ""),
            "Datasheet URL": p.get("datasheet_url", ""),
            "Source Owner": "Tavily research (pending review)",
            "Status": p.get("status", "Candidate"),
            "Notes": p.get("notes", ""),
        })
        r += 1

    # 3. Sheet 08 - parameters.
    param_hmap = _header_map(ws_param, param_hrow)
    _clear_data_rows(ws_param, param_hrow)
    r = param_hrow + 1
    for prm in parameters:
        mid = prm.get("metric_id", "")
        mt, mp = match_defaults.get(mid, ("", ""))
        if mt or mp:
            summary["match_defaults_reused"] += 1
        _write_row(ws_param, r, param_hmap, {
            "Product ID": prm["product_id"],
            "Product Model": prm.get("model", ""),
            "Product Family ID": prm.get("family_id", ""),
            "Parameter ID": mid,
            "Parameter Name": prm.get("parameter_name", ""),
            "Min Value": prm.get("min_value"),
            "Max Value": prm.get("max_value"),
            "Supported Value / Text": prm.get("supported_value", ""),
            "Unit": prm.get("unit", ""),
            "Match Type": mt,
            "Match Priority": mp,
            "Evidence Source / Datasheet URL": url_by_pid.get(prm["product_id"], ""),
            "Notes": prm.get("notes", ""),
        })
        r += 1

    wb.save(master_path)
    return summary


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Backfill master data package")
    parser.add_argument("--catalog", default=str(CATALOG_JSON),
                        help="Path to step0_product_catalog.json")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    s = backfill(Path(args.catalog), dry_run=args.dry_run)
    print("\n=== Backfill master data package ===")
    print(f"Master : {s['master']}")
    print(f"Products -> sheet 07 : {s['products_written']}")
    print(f"Parameters -> sheet 08: {s['parameters_written']}")
    if not s["dry_run"]:
        print(f"Match defaults reused : {s['match_defaults_reused']}")
        print(f"Backup saved to       : {s['backup']}")
    else:
        print("(dry run - no changes written)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
