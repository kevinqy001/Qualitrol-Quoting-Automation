"""Export a generated BOQ to an .xlsx that mirrors the customer BOQ template.

The reference template (``Gemba Samples/2/Templetes/MEA Example BOQ.xlsx``) is a
real project quote. Its transferable structure is:

    Row 1 (merged)      project / lot title
    Row 3               column headers:  DETAILS | ITEM | QTY
    grouped sections    "Desc.: <panel / scope>" header, then item rows

We reproduce that field layout and styling from the pipeline's BOQ output rather
than filling the project-specific master (which carries hardcoded bays, panels
and merged cells). Output columns:

    DETAILS | ITEM | QTY | UNIT | REVIEW

DETAILS  = line description + quantity basis / related assets
ITEM     = product model (falls back to product code)
QTY/UNIT = quantity and unit
REVIEW   = review status (Draft / Needs Review)

Open clarification questions are appended as their own section, matching the
docx generator's behaviour.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brand-ish flat colors (no gradients); kept subtle so the sheet prints cleanly.
_TITLE_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")
_REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADERS = ["DETAILS", "ITEM", "QTY", "UNIT", "REVIEW"]
_COL_WIDTHS = [52, 34, 10, 10, 18]


def _scenario_name_map(boq: dict) -> dict[str, str]:
    out: dict[str, str] = {}
    for det in boq.get("detectedScenarios", []) or []:
        sid = str(det.get("scenario_id", "")).strip()
        if sid:
            out[sid] = str(det.get("scenario", "")).strip() or sid
    return out


def _line_details(item: dict) -> str:
    """DETAILS cell: description plus the quantity basis / related assets."""
    parts: list[str] = []
    desc = (item.get("description") or "").strip()
    if desc:
        parts.append(desc)
    params = item.get("technicalParams") or {}
    basis = (params.get("basis") or "").strip()
    related = params.get("related")
    if related:
        related_str = ", ".join(related) if isinstance(related, list) else str(related)
        parts.append(f"Related: {related_str}")
    if basis:
        parts.append(f"Basis: {basis}")
    return "\n".join(parts) if parts else (item.get("productCode") or "")


def _group_key(item: dict, names: dict[str, str]) -> str:
    params = item.get("technicalParams") or {}
    sid = str(params.get("scenario", "")).strip()
    if sid:
        return names.get(sid, sid)
    return "BOQ Items"


def generate_boq_xlsx(
    boq: dict,
    out_path: Path,
    template_path: Optional[Path] = None,  # accepted for API symmetry; not required
) -> Path:
    """Write ``boq`` (the frontend extraction/BOQ dict) to a template-style xlsx."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    line_items = boq.get("lineItems") or []
    case_ref = boq.get("caseReference") or boq.get("boqId") or "DRAFT"
    names = _scenario_name_map(boq)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    for idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    n_cols = len(_HEADERS)
    last_col = get_column_letter(n_cols)

    # --- Title (merged across all columns) --------------------------------- #
    ws.merge_cells(f"A1:{last_col}1")
    title = ws["A1"]
    title.value = f"Qualitrol Monitoring BOQ — {case_ref}"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = _TITLE_FILL
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    summary = (boq.get("extractionSummary") or "").strip()
    if summary:
        ws.merge_cells(f"A2:{last_col}2")
        sub = ws["A2"]
        sub.value = summary
        sub.font = Font(italic=True, size=9, color="595959")
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[2].height = 22

    # --- Header row -------------------------------------------------------- #
    header_row = 4
    for col, label in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(
            horizontal="center" if col >= 3 else "left", vertical="center"
        )

    row = header_row + 1

    # --- Grouped BOQ line items -------------------------------------------- #
    if line_items:
        grouped: dict[str, list[dict]] = {}
        order: list[str] = []
        for item in line_items:
            key = _group_key(item, names)
            if key not in grouped:
                grouped[key] = []
                order.append(key)
            grouped[key].append(item)

        for key in order:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
            sec = ws.cell(row=row, column=1, value=f"Desc.: {key}")
            sec.font = Font(bold=True, size=10, color="1F3A5F")
            sec.fill = _SECTION_FILL
            sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).border = _BORDER
            row += 1

            for item in grouped[key]:
                review = (item.get("review_status") or "").strip()
                values = [
                    _line_details(item),
                    item.get("product_model") or item.get("productCode") or "",
                    item.get("quantity"),
                    item.get("unit") or "",
                    review,
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = _BORDER
                    if col == 1:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    elif col == 2:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.font = Font(size=10, bold=True)
                    elif col == 3:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                    elif col == 4:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                        if review and review.lower() != "draft":
                            cell.fill = _REVIEW_FILL
                row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        empty = ws.cell(row=row, column=1, value="No Qualitrol product lines detected.")
        empty.alignment = Alignment(horizontal="center", vertical="center")
        empty.font = Font(italic=True, color="999999")
        row += 1

    # --- Open clarification questions -------------------------------------- #
    questions = boq.get("missingInfoQuestions") or []
    if questions:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        head = ws.cell(row=row, column=1, value="Open Clarification Questions (resolve before quoting)")
        head.font = Font(bold=True, size=10, color="9C5700")
        head.fill = _REVIEW_FILL
        head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, n_cols + 1):
            ws.cell(row=row, column=col).border = _BORDER
        row += 1

        for col, label in enumerate(["QUESTION", "OWNER", "PRIORITY", "", ""], start=1):
            if not label:
                continue
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(bold=True, size=9)
            cell.fill = _HEADER_FILL
            cell.border = _BORDER
        row += 1

        for q in questions:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            qcell = ws.cell(row=row, column=1, value=q.get("question") or q.get("missing_item") or "")
            qcell.alignment = Alignment(vertical="top", wrap_text=True)
            qcell.border = _BORDER
            ws.cell(row=row, column=2).border = _BORDER
            owner = ws.cell(row=row, column=3, value=q.get("owner") or "")
            owner.alignment = Alignment(vertical="top", wrap_text=True)
            owner.border = _BORDER
            prio = ws.cell(row=row, column=4, value=q.get("priority") or "")
            prio.alignment = Alignment(horizontal="center", vertical="top")
            prio.border = _BORDER
            ws.cell(row=row, column=5).border = _BORDER
            row += 1

    ws.freeze_panes = "A5"
    wb.save(str(out_path))
    return out_path
