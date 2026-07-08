"""Copy today's KB additions from the main data package into the Shared review
copy (Preparation/..._Shared.xlsx) so business colleagues can review them.

APPEND-ONLY: only appends the rows added today (identified by ID / source
marker); never edits/deletes existing Shared rows; dedups; backs up the Shared
workbook first.
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
MAIN = REPO / "Qualitrol_BOQ_Matching_Data_Package.xlsx"
SHARED = REPO / "Preparation" / "Qualitrol_BOQ_Matching_Data_Package - Shared.xlsx"

ADDED_FAMILY_IDS = {"PF_DAU_REC", "PF_MON_PANEL", "PF_NET_SEC", "PF_TIMING",
                    "PF_SW_LIC", "PF_SERVICES", "PF_PANEL_ACC"}
ADDED_QR_IDS = {"QR_DAU_BAY_001", "QR_PANEL_001", "QR_PER_PANEL_001",
                "QR_PER_SYSTEM_001", "QR_PER_DAU_001", "QR_SERVICES_001"}
ADDED_PRODUCT_SOURCES = {"BOQ reverse-extraction", "TAQA MEA ruleset"}


def _header_idx(rows, first_col):
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip().lower() == first_col.lower():
            return i
    raise RuntimeError(f"header {first_col!r} not found")


def _added_rows(main_rows, header_i, predicate):
    return [list(r) for r in main_rows[header_i + 1:]
            if r and any(c is not None for c in r) and predicate(r)]


def _existing_keys(shared_rows, header_i, col):
    return {str(r[col]).strip().lower() for r in shared_rows[header_i + 1:]
            if r and len(r) > col and r[col] is not None}


def main():
    wb_main = openpyxl.load_workbook(MAIN, read_only=True, data_only=True)
    wb_shared = openpyxl.load_workbook(SHARED)

    plan = [
        # sheet, header first-col, id-col, predicate(row)->bool, dedup-col
        ("06_Product_Family_Master", "Product Family ID", 0,
         lambda r: str(r[0]).strip() in ADDED_FAMILY_IDS, 0),
        ("07_Product_Master_Template", "Product ID", 0,
         lambda r: len(r) > 13 and str(r[13]).strip() in ADDED_PRODUCT_SOURCES, 1),
        ("09_Quantity_Rules", "Quantity Rule ID", 0,
         lambda r: str(r[0]).strip() in ADDED_QR_IDS, 0),
        ("10_Compatibility_Rules", "Rule ID", 0,
         lambda r: str(r[0]).strip().startswith(("CR_MEA_", "CR_BAY_")), 0),
    ]

    report = {}
    for sheet, first_col, id_col, pred, dedup_col in plan:
        main_rows = list(wb_main[sheet].iter_rows(values_only=True))
        hi_m = _header_idx(main_rows, first_col)
        rows_to_add = _added_rows(main_rows, hi_m, pred)

        ws_s = wb_shared[sheet]
        shared_rows = list(ws_s.iter_rows(values_only=True))
        hi_s = _header_idx(shared_rows, first_col)
        existing = _existing_keys(shared_rows, hi_s, dedup_col)

        added = skipped = 0
        for row in rows_to_add:
            key = str(row[dedup_col]).strip().lower() if len(row) > dedup_col and row[dedup_col] is not None else ""
            if key and key in existing:
                skipped += 1
                continue
            ws_s.append(row)
            existing.add(key)
            added += 1
        report[sheet] = (added, skipped)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = SHARED.with_name(f"{SHARED.stem}.backup_{stamp}.xlsx")
    shutil.copyfile(SHARED, backup)
    wb_shared.save(SHARED)
    print("Shared backup:", backup.name)
    for s, (a, k) in report.items():
        print(f"  {s}: added {a}, skipped {k}")


if __name__ == "__main__":
    main()
