import openpyxl, sys

def dump(path, out):
    wb = openpyxl.load_workbook(path, data_only=True)
    with open(out, 'w', encoding='utf-8') as f:
        for ws in wb.worksheets:
            f.write(f"\n{'='*100}\nSHEET: {ws.title}  (dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})\n{'='*100}\n")
            for row in ws.iter_rows(values_only=True):
                # skip fully empty rows
                if all(c is None or (isinstance(c,str) and c.strip()=='') for c in row):
                    continue
                cells = []
                for c in row:
                    if c is None:
                        cells.append('')
                    else:
                        cells.append(str(c).replace('\n',' / '))
                f.write(' | '.join(cells) + '\n')
    print("wrote", out)

files = [
    (r'C:\Users\yqiao2\Ralliant\QTC-Digital-Enablement - Qualitrol x Foundry - Projects Quoting Hack\User Testing\775368\BOQ-WEB-48DB5A88 - 775368.xlsx', 'dump_775368.txt'),
    (r'C:\Users\yqiao2\Ralliant\QTC-Digital-Enablement - Qualitrol x Foundry - Projects Quoting Hack\User Testing\776060\BOQ-WEB-365D2B5A - 776060.xlsx', 'dump_776060.txt'),
]
for p,o in files:
    dump(p,o)
