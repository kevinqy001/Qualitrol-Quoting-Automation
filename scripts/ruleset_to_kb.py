"""Append the TAQA/ADNOC MEA configuration ruleset into the data package.

Source: 'Ruleset for TAQA ADNOC MEA Region - IP.xlsx' (All Options sheet).
APPEND-ONLY: adds new Compatibility Rules (device selection, accessory ratios,
panel BoM bundles) and the ruleset's TAQA-approved chassis variants as products.
Never edits/deletes existing rows; dedups by Rule ID / Product Model; backs up
the workbook first.
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
PKG = REPO / "Qualitrol_BOQ_Matching_Data_Package.xlsx"
SRC = "TAQA/ADNOC MEA ruleset (Ruleset for TAQA ADNOC MEA Region - IP.xlsx); candidate — review before use."

CR_SHEET, PROD_SHEET = "10_Compatibility_Rules", "07_Product_Master_Template"

# Rule ID, Rule Type, Scenario ID, Asset Type, Condition / Trigger,
# Recommended Action, Severity, Notes
RULES = [
    ("CR_MEA_01", "Selection", "FMS_001", "Feeder / bay",
     "FMS monitoring required on ALL feeders/bays ('Everywhere').",
     "Provide FMS on every feeder; minimum device IDM+ 9A/32D (3U). Increase analogue/digital channels as necessary.",
     "Medium", SRC),
    ("CR_MEA_02", "Selection", "PQ_CLASSA_001", "BS / BC / TX / Gen / 3rd-party feeder",
     "PQM required on Bus Section (BS), Bus Coupler (BC), Transformer (TX) feeders, 3rd-party feeders and Gen feeders.",
     "Provide PQM on these feeders using IDM+ 18A/64D or IDM+ 27A/96D (6U chassis).",
     "Medium", SRC),
    ("CR_MEA_03", "Selection", "PMU_001", "OHL / IC / Gen / 3rd-party feeder",
     "PMU required on OHL feeders, IC feeder, Gen feeders and 3rd-party feeders; NOT on BC, NOT on BS.",
     "Provide PMU on these feeders using IDM+ 36A/128D (6U). Do NOT place PMU on bus coupler / bus section feeders.",
     "Medium", SRC),
    ("CR_MEA_04", "Selection", "All", "DAU chassis",
     "TAQA-approved DAU chassis options.",
     "Approved chassis: 3U 9A/32D (FMS/PQM/PMU); 6U 18A/64D (FMS/PQM or FMS/PMU). Comms = RJ45. Increase A/D as necessary.",
     "Low", SRC),
    ("CR_MEA_05", "Quantity", "All", "Panel",
     "Panel device capacity limit.",
     "Max 4× 3U devices OR 2× 6U devices per panel; provide Option 1 & Option 2 when it affects panel count.",
     "Medium", SRC),
    ("CR_MEA_06", "Quantity", "PMU_001", "GPS timing",
     "GPS master / antenna sizing.",
     "1 GPS Master per 12 DAU; 2 antennas incl. 100m cable, amplifier, mounting kit & surge arrestor.",
     "Medium", SRC),
    ("CR_MEA_07", "Quantity", "COMM_SCADA_001", "Ethernet switch",
     "Monitoring-LAN port sizing.",
     "16-port Ethernet switch; allow 1 port per DAU; offer redundancy.",
     "Low", SRC),
    ("CR_MEA_08", "Quantity", "PMU_001", "EPG licence",
     "EPG licence count per device.",
     "4 EPG licenses per device (FMS/PMU chassis); N/A for PQM 6U chassis.",
     "Medium", SRC),
    ("CR_MEA_09", "Quantity", "All", "Test switch",
     "Test switch allocation.",
     "Provide test switches per device.",
     "Low", SRC),
    ("CR_MEA_10", "Quantity", "All", "Annunciator",
     "Annunciator allocation.",
     "Provide one annunciator per panel.",
     "Low", SRC),
    ("CR_MEA_11", "Bundle", "COMM_SCADA_001", "LEV panel",
     "FMS/PQM LEV panel standard contents.",
     "Include: Industrial Rack Mounted PC, Trend Micro Antivirus, Trellix Whitelisting, Acronis Backup, iQ+ Software, "
     "16 Port RJ45 4-port fiber L2 switch, 16 Port RJ45 4-port fiber L3 switch, Firewall, Printer, Annunciator.",
     "Medium", SRC),
    ("CR_MEA_12", "Bundle", "PMU_001;WAMS_001", "PDC panel",
     "PDC panel standard contents.",
     "Same as FMS/PQM LEV panel PLUS: Redundant Industrial Rack Mounted PC, eSPDC Software, total number of EPG Licenses, KVM.",
     "Medium", SRC),
]

# TAQA-approved chassis variants + eSPDC that the ruleset names.
# Model, Family ID, Family, Scenarios, Asset, Description, Default QR
_D = "PF_DAU_REC", "Multi-function DAU / Recorder (IDM+ / Informa)"
_S = "PF_SW_LIC", "Monitoring Software & Licences"
PRODUCTS = [
    ("IDM+ 18A/64D", *_D, "FMS_001;PQ_CLASSA_001", "Bay / Feeder",
     "6U DAU chassis, 18 analogue / 64 digital channels; TAQA-approved for FMS/PQM (and FMS/PMU).", "QR_DAU_BAY_001"),
    ("IDM+ 27A/96D", *_D, "PQ_CLASSA_001", "Bay / Feeder",
     "6U DAU chassis, 27 analogue / 96 digital channels; TAQA-approved for PQM.", "QR_DAU_BAY_001"),
    ("IDM+ 36A/128D", *_D, "PMU_001", "Bay / Feeder",
     "6U DAU chassis, 36 analogue / 128 digital channels; TAQA-approved for PMU on OHL/IC/Gen feeders.", "QR_DAU_BAY_001"),
    ("eSPDC Software", *_S, "PMU_001;WAMS_001", "System",
     "eSPDC phasor data concentrator software for the PDC panel.", "QR_PER_SYSTEM_001"),
]


def _header_row(rows, first_col):
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip().lower() == first_col.lower():
            return i
    raise RuntimeError(f"header {first_col!r} not found")


def main():
    wb = openpyxl.load_workbook(PKG)
    added = {"rules": 0, "products": 0}
    skipped = {"rules": 0, "products": 0}

    cr = wb[CR_SHEET]; crr = list(cr.iter_rows(values_only=True))
    ch = _header_row(crr, "Rule ID")
    rule_ids = {str(r[0]).strip() for r in crr[ch + 1:] if r and r[0]}
    for row in RULES:
        if row[0] in rule_ids:
            skipped["rules"] += 1; continue
        cr.append(list(row)); added["rules"] += 1

    prod = wb[PROD_SHEET]; pr = list(prod.iter_rows(values_only=True))
    ph = _header_row(pr, "Product ID")
    models = {str(r[1]).strip().lower() for r in pr[ph + 1:] if r and len(r) > 1 and r[1]}
    idx = 0
    for model, fid, fname, scen, asset, desc, qr in PRODUCTS:
        if model.strip().lower() in models:
            skipped["products"] += 1; continue
        idx += 1
        pid = f"PROD_MEA_{fid}_{idx:02d}"
        prod.append([pid, model, fid, fname, scen, asset, desc, "", "", qr,
                     "", "", "", "TAQA MEA ruleset", "candidate", SRC])
        added["products"] += 1

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = PKG.with_name(f"Qualitrol_BOQ_Matching_Data_Package.backup_{stamp}.xlsx")
    shutil.copyfile(PKG, backup)
    wb.save(PKG)
    print("Backup:", backup.name)
    print("ADDED:", added, "| SKIPPED:", skipped)


if __name__ == "__main__":
    main()
