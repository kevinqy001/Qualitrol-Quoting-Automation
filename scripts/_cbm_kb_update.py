"""CBM real-case KB update (2026-07, CBM KB folder: 14 transformer CBM projects).

Business context that drives this update
----------------------------------------
CBM / transformer condition-monitoring projects usually have **no SLD / drawing**
at the BOQ quoting stage. Step 1 & Step 2 must therefore be driven by the project
tender / technical specification, and product quantities are **counted from the
monitored-parameter list** stated in the document (temperature points, analog
inputs, RTD, CT, tap position, DGA gas count, bushings, PD sensors) — normally in
the "Online Monitoring" / "TMS-QTMS" / "Technical Particulars" section — NOT from
a drawing.

Sources: CBM KB real cases (South Africa LZ264537, Brazil COPEL LZ264403, Greece
AVAX LZ264341, USA Hitachi LZ264435, Norway Green Mountain LZ264392, Kuwait BESS
LZ263476, Canada Rose Valley LZ250523, Romania LZ264232/LZ263014, UAE Transco
LZ264194, Indonesia PLN GSUT LZ250378, Saudi GE-Sadawi LZ264438 / Jeddah LZ248337)
- the shared "CBM SYSTEM & SENSOR INPUT OVERVIEW PER TRANSFORMER BANK" configs,
the "SS Quote Config Form V4.0" (Parameters-Solution / CBM Package) and the
tender specs (Rose Valley MPT DRS, ABB Indonesia "Online monitor trafo").

The script is append-only for NEW rows and calibrates a small set of EXISTING
rows (transformer scenarios' Drawing Dependency + QTMS/DGA products' scenario
links). Re-running is safe: rows whose key already exists are skipped. A
timestamped backup is written first.
"""
import os
import shutil
import datetime
import openpyxl

XLSX = "Qualitrol_BOQ_Matching_Data_Package.xlsx"
STAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
REVIEW = (
    f"Added {STAMP[:8]} from CBM KB real-case scan "
    "(14 transformer CBM projects). REVIEW before use."
)

# --------------------------------------------------------------------------- #
# Sheet helpers
# --------------------------------------------------------------------------- #
def header_row_idx(ws, first_col):
    target = first_col.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip().lower() == target:
                return cell.row
    raise ValueError(f"header {first_col!r} not found in {ws.title}")


def header_list(ws, first_col):
    r = header_row_idx(ws, first_col)
    hdr = [("" if c.value is None else str(c.value).strip()) for c in ws[r]]
    while hdr and hdr[-1] == "":
        hdr.pop()
    return hdr, r


def col_index(hdr, name):
    return hdr.index(name) + 1  # 1-based


def existing_ids(ws, first_col, id_col_name=None):
    hdr, hr = header_list(ws, first_col)
    idc = col_index(hdr, id_col_name or first_col)
    out = set()
    for row in ws.iter_rows(min_row=hr + 1):
        v = row[idc - 1].value
        if v is not None and str(v).strip():
            out.add(str(v).strip())
    return out


def row_from(hdr, d):
    unknown = set(d) - set(hdr)
    if unknown:
        raise KeyError(f"unknown headers {unknown} for {hdr[0]!r}")
    return [d.get(h, "") for h in hdr]


def append_dicts(ws, first_col, dict_rows, id_key):
    hdr, _ = header_list(ws, first_col)
    have = existing_ids(ws, first_col, id_key)
    n = 0
    for d in dict_rows:
        if str(d.get(id_key, "")).strip() in have:
            continue
        ws.append(row_from(hdr, d))
        n += 1
    return n


# --------------------------------------------------------------------------- #
# 03 - NEW integrated CBM scenario  +  calibrate existing transformer scenarios
# --------------------------------------------------------------------------- #
SCENARIOS = [
    {
        "Scenario ID": "TR_CBM_001",
        "Scenario Category": "Transformer",
        "Application Scenario": "Integrated transformer condition monitoring system (QTMS)",
        "Asset Type": "Oil-filled power transformer (transformer bank)",
        "Typical Metrics / Requirements": (
            "Top/bottom oil temperature; winding & hot-spot temperature; load current; "
            "cooling (fan/pump) status & current; oil level (main + LTC tank); tank pressure; "
            "DGA gas concentrations + moisture; bushing capacitance/tan delta; partial discharge; "
            "OLTC tap position / drive-motor current; direct fibre-optic winding temperature; environment"
        ),
        "Common Evidence Keywords / Synonyms": (
            "online monitoring; on-line monitoring; transformer monitoring system; condition monitoring; CBM; "
            "QTMS; total transformer monitoring; real-time monitoring system; transformer life management monitoring; "
            "在线监测; 智能温度监测; 变压器在线监测; 本体在线监测"
        ),
        "Related Product Families": (
            "PF_TR_CBM; PF_TR_TEMP; PF_DGA; PF_BUSHING; PF_TR_PD; PF_OLTC; PF_AUX_SENSOR; PF_SOFTWARE"
        ),
        "Quantity Basis": (
            "1 integrated QTMS system per transformer bank. Internal modules/sensors are counted from the "
            "monitored-parameter list in the project tender/technical spec (temperature points, analog inputs, "
            "RTD, CT, tap position, DGA gas count, bushings, PD sensors) - NOT from an SLD. "
            "Analog Input module = ceil(analog points/8); Digital Input = ceil(digital points/14); "
            "Fibre-Optic card = 4/6/8 ch; Output Relay = ceil(outputs/8)."
        ),
        "Drawing Dependency": (
            "None at BOQ stage - CBM transformer projects usually have no SLD when quoting; "
            "derive scope and quantity from the tender / project document (monitored-parameter list)."
        ),
        "Requirement Output Fields": (
            "asset_tag; analog_input_count; rtd_count; ct_count; tap_position; dga_gas_count; "
            "bushing_count; pd_sensor_count; fo_channel_count"
        ),
        "Review Notes": REVIEW + " Integrated umbrella scenario for transformer CBM (per-transformer-bank QTMS config).",
    },
]

# Existing transformer scenarios -> set Drawing Dependency to reflect "no SLD at
# BOQ; count from project document", and append a calibration note.
SCENARIO_DRAWING_CAL = {
    "TR_DGA_001": "None at BOQ stage - count from transformer list in the project document.",
    "TR_TEMP_001": "None at BOQ stage - temperature/analog points counted from the project spec, not a drawing.",
    "TR_BUSH_001": "None at BOQ stage - bushing count taken from the project spec / nameplate, not a drawing.",
    "TR_AUX_001": "None at BOQ stage - accessory list taken from the project spec, not a drawing.",
    "TAPCHG_001": "None at BOQ stage - OLTC/tap monitoring scope taken from the project spec, not a drawing.",
    "TR_PD_001": "Optional - for transformers, PD sensor count from transformer count/design in the spec (no SLD needed at BOQ).",
}
CAL_NOTE = (
    " [Calibrated 2026-07 CBM KB scan: CBM transformer projects have no SLD at BOQ; "
    "quantity derived from the project document monitored-parameter list.]"
)


def calibrate_scenarios(ws):
    hdr, hr = header_list(ws, "Scenario ID")
    c_id = col_index(hdr, "Scenario ID")
    c_draw = col_index(hdr, "Drawing Dependency")
    c_note = col_index(hdr, "Review Notes")
    n = 0
    for row in ws.iter_rows(min_row=hr + 1):
        sid = row[c_id - 1].value
        sid = str(sid).strip() if sid is not None else ""
        if sid in SCENARIO_DRAWING_CAL:
            row[c_draw - 1].value = SCENARIO_DRAWING_CAL[sid]
            cur = row[c_note - 1].value or ""
            if "Calibrated 2026-07 CBM KB" not in str(cur):
                row[c_note - 1].value = (str(cur) + CAL_NOTE).strip()
            n += 1
    return n


# --------------------------------------------------------------------------- #
# 04 - Metrics
# --------------------------------------------------------------------------- #
def metric(mid, name, syn, unit, dtype, applies, used, examples=""):
    return {
        "Metric ID": mid, "Standard Metric Name": name, "Synonyms / Raw Terms": syn,
        "Standard Unit": unit, "Data Type": dtype, "Applies To": applies, "Used For": used,
        "Example Values": examples, "Required for Matching": "No", "Normalization Notes": REVIEW,
    }


METRICS = [
    metric("MET_HOTSPOT_TEMP", "Winding Hot-Spot Temperature",
           "hot spot; hotspot; winding hotspot; hot-spot temperature", "degC", "number/range",
           "Transformer", "Winding hot-spot temperature monitoring / fan staging"),
    metric("MET_LOAD_CURRENT", "Load Current",
           "load current; transformer load; loading; load monitoring", "A", "number",
           "Transformer", "Load / winding-temperature-simulation monitoring"),
    metric("MET_COOLING_STATUS", "Cooling System Status",
           "cooling status; fan status; pump status; cooling system monitoring; fan/pump current", "text", "text",
           "Transformer", "Cooling (fan/pump) status & current monitoring"),
    metric("MET_RTD_COUNT", "RTD / Temperature Point Count",
           "number of RTD; RTD count; Pt100 count; temperature point count", "count", "integer",
           "Transformer", "Sizes QTMS Analog Input module (RTD card)"),
    metric("MET_ANALOG_INPUT_COUNT", "Analog Input Point Count",
           "analog input count; analog points; number of analog inputs; 4-20mA points", "count", "integer",
           "Transformer", "Sizes QTMS Analog Input module (8 inputs/module)"),
    metric("MET_DIGITAL_INPUT_COUNT", "Digital Input Point Count",
           "digital input count; digital points; status inputs; dry contact count", "count", "integer",
           "Transformer", "Sizes QTMS Digital Input module (14 inputs/module)"),
    metric("MET_FO_CHANNEL_COUNT", "Direct Winding Temperature Channel Count",
           "fibre optic channels; FO channels; direct winding temperature sensors; number of FO sensors", "count", "integer",
           "Transformer", "Sizes QTMS Fibre-Optic module (4/6/8 inputs) + fibre/OFT/TWP"),
    metric("MET_RELAY_OUTPUT_COUNT", "Relay Output Count",
           "relay outputs; alarm/trip outputs; number of relays", "count", "integer",
           "Transformer", "Sizes QTMS Output Relay module (8 outputs/module)"),
]

# --------------------------------------------------------------------------- #
# 06 - Family
# --------------------------------------------------------------------------- #
FAMILIES = [
    {
        "Product Line": "CBM",
        "Product Family ID": "PF_TR_CBM",
        "Product Family": "Integrated Transformer Monitoring System (QTMS)",
        "Applicable Scenario IDs": "TR_CBM_001",
        "Primary Asset Type": "Oil-filled power transformer (bank)",
        "Typical Capabilities": (
            "Modular QTMS controller (3U/6U): analog input (RTD/CT/tap-position/4-20mA), digital input, "
            "fibre-optic direct winding temperature, bushing, output relay, power supply, display/remote display; "
            "integrates DGA (Serveron TM1/TM3/TM8), QPDM partial discharge, power meters, and SmartSUB APM software."
        ),
        "Default Quantity Rule ID": "QR_TR_CBM_SYS_001",
        "Dependencies / Required Inputs": (
            "Transformer bank count; monitored-parameter list from tender/spec (temperature points, analog inputs, "
            "tap position, DGA gas count, bushings, PD sensors); voltage level (3U vs 6U chassis)."
        ),
        "Commercial / Engineering Notes": REVIEW + " Umbrella family for the per-transformer-bank CBM configuration.",
    },
]

# --------------------------------------------------------------------------- #
# 07 - Products (net-new)   +   calibration of existing QTMS/DGA product scenarios
# --------------------------------------------------------------------------- #
def prod(pid, model, fid, fam, scen, asset, desc, rule="", proto=""):
    return {
        "Product ID": pid, "Product Model": model, "Product Family ID": fid, "Product Family": fam,
        "Applicable Scenario IDs": scen, "Primary Asset Type": asset, "Product Description": desc,
        "Supported Standards": "", "Communication Protocols": proto, "Default Quantity Rule ID": rule,
        "Required Accessories": "", "Optional Accessories": "", "Datasheet URL": "",
        "Source Owner": "CBM KB scan 2026-07", "Status": "Review", "Notes": REVIEW,
    }


CBM = ("PF_TR_CBM", "Integrated Transformer Monitoring System (QTMS)")
PRODUCTS = [
    prod("CBM_QTMS_BASE_6U", "QTMS Base Module 6U", *CBM, "TR_CBM_001", "Power transformer",
         "QTMS 6U base panel: Chassis + CPU + Display, ~11 module slots; for high module count (400/230kV banks with FO + PD + bushing).",
         rule="QR_TR_CBM_SYS_001", proto="Modbus; IEC 61850; DNP3"),
    prod("CBM_QTMS_BASE_3U", "QTMS Base Module 3U", *CBM, "TR_CBM_001", "Power transformer",
         "QTMS 3U base panel: Chassis + CPU + Display, ~5 module slots; for smaller banks (150/132kV).",
         rule="QR_TR_CBM_SYS_001", proto="Modbus; IEC 61850; DNP3"),
    prod("CBM_RTD_PT100", "Universal RTD Pt100 (103-044 series)", *CBM, "TR_CBM_001;TR_TEMP_001", "Transformer",
         "One-piece universal RTD Pt100 (oil/ambient/cooling/OLTC temperature) to QTMS AI RTD card.", rule="QR_TR_CBM_AI_001"),
    prod("CBM_CLAMP_CT", "Clamp-On CT (TRA-017 series)", *CBM, "TR_CBM_001", "Transformer",
         "Clamp-on CT (20ft leads) for winding-temperature simulation / load / fan / pump / OLTC-drive current to QTMS AI CT card.", rule="QR_TR_CBM_AI_001"),
    prod("CBM_TAP_INPUT", "Tap-Position Input (4-20mA)", *CBM, "TR_CBM_001;TAPCHG_001", "On-load tap changer",
         "QTMS AI 4-20mA input for OLTC tap position (sensor customer-supplied).", rule="QR_TR_CBM_AI_001"),
    prod("CBM_POWER_METER", "Power Meter (OSP, Modbus)", *CBM, "TR_CBM_001", "Transformer",
         "OSP power meter with clamp-on CT integrated to QTMS via Modbus.", rule="QR_TR_CBM_SYS_001", proto="RS-485 Modbus"),
    prod("CBM_QBM_ADAPT", "Bushing Adaptor with Sensor", "PF_BUSHING", "Bushing Monitor",
         "TR_CBM_001;TR_BUSH_001", "Transformer bushing",
         "Bushing tap adaptor + sensor for capacitance / tan-delta monitoring (one per monitored bushing).", rule="QR_TR_CBM_BUSH_001"),
    prod("CBM_QBM_TPCABLE", "Bushing Twisted-Pair Cable (25m)", "PF_BUSHING", "Bushing Monitor",
         "TR_CBM_001;TR_BUSH_001", "Transformer bushing",
         "Twisted-pair cable from bushing adaptor to bushing monitor (one per monitored bushing).", rule="QR_TR_CBM_BUSH_001"),
    prod("CBM_FO_INTERNAL", "Internal Fibre-Optic Winding Temp Probe (T2S series)", *CBM, "TR_CBM_001;TR_TEMP_001", "Transformer",
         "Internal GaAs fibre-optic probe (up to ~12m) for direct winding temperature, factory-installed.", rule="QR_TR_CBM_FO_001"),
    prod("CBM_FO_EXTERNAL", "External Fibre (Ext-3MP series) + OFT Feedthrough", *CBM, "TR_CBM_001;TR_TEMP_001", "Transformer",
         "External fibre (up to ~10m) and OFT tank-wall feedthrough for each direct-winding-temperature channel.", rule="QR_TR_CBM_FO_001"),
    prod("CBM_NXP611", "NXP-611 (OFT/TWP factory leak-test service)", *CBM, "TR_CBM_001;TR_TEMP_001", "Transformer",
         "Factory service to leak-test OFT feedthroughs in the TWP plate (one per transformer with FO winding temp).", rule="QR_TR_CBM_FO_001"),
    prod("CBM_TM_MOIST", "Serveron Moisture Sensor", "PF_DGA", "Online DGA Monitor",
         "TR_CBM_001;TR_DGA_001", "Transformer",
         "Moisture-in-oil sensor for Serveron TM1/TM3/TM8 DGA monitor (one per DGA monitor).", rule="QR_TR_CBM_DGA_001"),
    prod("CBM_TM_ACCESS", "Serveron DGA Accessory Kit", "PF_DGA", "Online DGA Monitor",
         "TR_CBM_001;TR_DGA_001", "Transformer",
         "DGA install accessories: mounting kit (stand/pad), steel tubing (8 tubes/unit), calibration gas / helium.", rule="QR_TR_CBM_DGA_001"),
]

# Existing products to link into the integrated scenario TR_CBM_001.
PRODUCT_SCEN_CAL = {
    "GMB_QTMS_BASE": "TR_CBM_001", "GMB_QTMS_AI": "TR_CBM_001", "GMB_QTMS_DI": "TR_CBM_001",
    "GMB_QTMS_FO": "TR_CBM_001", "GMB_QTMS_BM": "TR_CBM_001", "GMB_QTMS_RO": "TR_CBM_001",
    "GMB_QTMS_PSM": "TR_CBM_001", "GMB_QTMS_LLG": "TR_CBM_001", "GMB_QTMS_PRESS": "TR_CBM_001",
    "GMB_QTMS_FOPROBE": "TR_CBM_001", "GMB_QTMS_LTC": "TR_CBM_001",
    "PROD_PF_DGA_01": "TR_CBM_001", "PROD_PF_DGA_02": "TR_CBM_001", "PROD_PF_DGA_03": "TR_CBM_001",
}


def calibrate_products(ws):
    hdr, hr = header_list(ws, "Product ID")
    c_id = col_index(hdr, "Product ID")
    c_scen = col_index(hdr, "Applicable Scenario IDs")
    n = 0
    for row in ws.iter_rows(min_row=hr + 1):
        pid = row[c_id - 1].value
        pid = str(pid).strip() if pid is not None else ""
        if pid in PRODUCT_SCEN_CAL:
            add = PRODUCT_SCEN_CAL[pid]
            cur = str(row[c_scen - 1].value or "").strip()
            parts = [p.strip() for p in cur.replace(";", " ").split() if p.strip()]
            if add not in parts:
                row[c_scen - 1].value = (cur + ("; " if cur else "") + add).strip("; ").strip()
                n += 1
    return n


# --------------------------------------------------------------------------- #
# 09 - Quantity rules
# --------------------------------------------------------------------------- #
def qrule(rid, scen, fid, fam, basis, desc, count_field="", need_drawing="No",
          need_asset="No", example="", assumption=""):
    return {
        "Quantity Rule ID": rid, "Scenario ID": scen, "Product Family ID": fid, "Product Family": fam,
        "Quantity Basis": basis, "Rule Description": desc, "Need Drawing": need_drawing,
        "Need Asset List": need_asset, "Count Field": count_field, "Example": example,
        "Assumption / Risk": (assumption + " " if assumption else "") + REVIEW,
    }


QRULES = [
    qrule("QR_TR_CBM_SYS_001", "TR_CBM_001", "PF_TR_CBM", CBM[1],
          "Transformer bank count",
          "1 integrated QTMS monitoring system per transformer bank. 6U chassis for high module count "
          "(e.g. 400/230kV banks with FO + PD + bushing); 3U for smaller (150/132kV).",
          count_field="transformer_count", example="2 x 230kV banks => 2 QTMS systems"),
    qrule("QR_TR_CBM_AI_001", "TR_CBM_001", "PF_TR_CBM", CBM[1],
          "Analog input point count",
          "Analog Input module (8 inputs each) = ceil(total analog points / 8). Analog points = RTD (oil/ambient/"
          "cooling/OLTC temp) + CT (winding-temp simulation, fan/pump, OLTC drive) + tap-position 4-20mA + spare. "
          "Counted from the tender monitored-parameter list, NOT from a drawing.",
          count_field="analog_input_count", example="21 analog points => 3 AI modules"),
    qrule("QR_TR_CBM_DI_001", "TR_CBM_001", "PF_TR_CBM", CBM[1],
          "Digital input point count",
          "Digital Input module (14 inputs each) = ceil(digital status points / 14). Digital points = cooling status, "
          "temperature/level/pressure alarms, OLTC switching signals.",
          count_field="digital_input_count", example="14 digital points => 1 DI module"),
    qrule("QR_TR_CBM_FO_001", "TR_CBM_001", "PF_TR_CBM", CBM[1],
          "Direct winding temperature channel count",
          "Fibre-Optic module sized by direct winding-temperature channel count (4/6/8-input card). Per channel: "
          "internal fibre + external fibre + OFT feedthrough; 1 TWP plate + 1 NXP-611 per transformer.",
          count_field="fo_channel_count", example="15 FO channels => 2 x 8-ch FO modules + 15 fibres/OFT + 1 TWP + 1 NXP-611"),
    qrule("QR_TR_CBM_RO_001", "TR_CBM_001", "PF_TR_CBM", CBM[1],
          "Relay output count",
          "Output Relay module (8 outputs each) = ceil(alarm/trip outputs / 8).",
          count_field="relay_output_count", example="16 outputs => 2 RO modules"),
    qrule("QR_TR_CBM_DGA_001", "TR_CBM_001;TR_DGA_001", "PF_DGA", "Online DGA Monitor",
          "DGA gas count per transformer",
          "1 DGA monitor per transformer, selected by required gas count: Serveron TM8 (8-9 gas) / TM3 (3 gas) / "
          "TM1 (H2 only). Add moisture sensor, mounting kit, steel tubing, calibration gas per monitor.",
          count_field="dga_gas_count", need_asset="Yes", example="9-gas DGA requirement => 1 x Serveron TM8"),
    qrule("QR_TR_CBM_BUSH_001", "TR_CBM_001;TR_BUSH_001", "PF_BUSHING", "Bushing Monitor",
          "Monitored bushing count",
          "Bushing adaptor+sensor and twisted-pair cable = monitored bushing count. One bushing-monitor host handles "
          "max 6 bushings; >6 (e.g. 9) needs a second host / two bushing-monitor types.",
          count_field="bushing_count", need_asset="Yes", example="9 bushings (6x132kV + 3x275kV) => 2 hosts"),
]

# --------------------------------------------------------------------------- #
# 05 - Synonyms
# --------------------------------------------------------------------------- #
SYNONYMS = [
    ("online monitoring", "TR_CBM_001", "", "Integrated online transformer monitoring", "Transformer", "High"),
    ("on-line monitoring", "TR_CBM_001", "", "Integrated online transformer monitoring", "Transformer", "High"),
    ("transformer monitoring system", "TR_CBM_001", "", "Integrated transformer monitoring system (QTMS)", "Transformer", "High"),
    ("total transformer monitoring", "TR_CBM_001", "", "Integrated (total) transformer monitoring", "Transformer", "High"),
    ("real time monitoring system", "TR_CBM_001", "", "Real-time transformer monitoring system", "Transformer", "Medium"),
    ("transformer life management monitoring", "TR_CBM_001", "", "Online transformer life management monitoring", "Transformer", "High"),
    ("在线监测", "TR_CBM_001", "", "Online monitoring (Chinese)", "Transformer", "High"),
    ("智能温度监测", "TR_TEMP_001", "", "Intelligent temperature monitoring (Chinese)", "Transformer", "High"),
    ("套管在线监测", "TR_BUSH_001", "", "Bushing online monitoring (Chinese)", "Transformer bushing", "High"),
    ("本体在线监测", "TR_CBM_001", "", "Transformer main-body online monitoring (Chinese)", "Transformer", "High"),
    ("QTMS Base Module", "TR_CBM_001", "", "QTMS modular base panel (3U/6U)", "Transformer", "High"),
    ("direct winding temperature", "TR_TEMP_001", "MET_FO_WINDING_TEMP", "Direct fibre-optic winding temperature", "Transformer", "High"),
    ("winding hotspot", "TR_TEMP_001", "MET_HOTSPOT_TEMP", "Winding hot-spot temperature", "Transformer", "High"),
    ("WTI", "TR_TEMP_001", "MET_WINDING_TEMP", "Winding temperature indicator", "Transformer", "Medium"),
    ("OTI", "TR_TEMP_001", "MET_OIL_TEMPERATURE", "Oil temperature indicator", "Transformer", "Medium"),
    ("RTD Pt100", "TR_TEMP_001", "MET_RTD_COUNT", "RTD Pt100 temperature sensor", "Transformer", "High"),
    ("Pt100", "TR_TEMP_001", "MET_RTD_COUNT", "Pt100 RTD sensor", "Transformer", "Medium"),
    ("clamp-on CT", "TR_CBM_001", "MET_LOAD_CURRENT", "Clamp-on CT (load/fan/pump/OLTC current)", "Transformer", "Medium"),
    ("clamp on CT", "TR_CBM_001", "MET_LOAD_CURRENT", "Clamp-on CT", "Transformer", "Medium"),
    ("OFT", "TR_TEMP_001", "MET_FO_CHANNEL_COUNT", "Optical feedthrough (fibre-optic winding temp)", "Transformer", "Medium"),
    ("TWP", "TR_TEMP_001", "MET_FO_CHANNEL_COUNT", "Tank wall plate for OFT feedthroughs", "Transformer", "Medium"),
    ("NXP-611", "TR_TEMP_001", "", "OFT/TWP factory leak-test service", "Transformer", "Low"),
    ("twisted pair cable", "TR_BUSH_001", "MET_BUSHING_COUNT", "Bushing monitor twisted-pair cable", "Transformer bushing", "Low"),
    ("Neoptix", "TR_TEMP_001", "MET_FO_WINDING_TEMP", "Neoptix fibre-optic (GaAs) winding temperature sensor", "Transformer", "Medium"),
    ("GaAs sensor", "TR_TEMP_001", "MET_FO_WINDING_TEMP", "Gallium-arsenide fibre-optic temperature sensor", "Transformer", "Low"),
    ("marshalling kiosk", "TR_CBM_001", "", "Marshalling kiosk / control cabinet housing the monitor", "Transformer", "Low"),
    ("cooling system monitoring", "TR_CBM_001", "MET_COOLING_STATUS", "Cooling (fan/pump) monitoring", "Transformer", "Medium"),
    ("load monitoring", "TR_CBM_001", "MET_LOAD_CURRENT", "Transformer load monitoring", "Transformer", "Medium"),
    ("tap position input", "TAPCHG_001", "MET_TAP_POSITION", "OLTC tap-position 4-20mA input", "On-load tap changer", "High"),
    ("DETC", "TAPCHG_001", "", "De-energised / off-circuit tap changer (no online tap monitoring)", "Tap changer", "Medium"),
    ("power meter", "TR_CBM_001", "", "Power meter integrated to QTMS (OSP Modbus)", "Transformer", "Low"),
    ("moisture in oil", "TR_DGA_001", "MET_MOISTURE_OIL", "Moisture-in-oil (DGA monitor)", "Transformer", "Medium"),
    ("SmartSUB", "SUB_SOFT_001", "", "SmartSUB APM software (1 licence per asset)", "Transformer", "Medium"),
]


def append_synonyms(ws):
    hdr, hr = header_list(ws, "Raw Term / Phrase")
    have = existing_ids(ws, "Raw Term / Phrase")
    n = 0
    for term, sid, mid, meaning, ctx, prio in SYNONYMS:
        if term in have:
            continue
        d = {"Raw Term / Phrase": term, "Mapped Scenario ID": sid, "Mapped Metric ID": mid,
             "Mapped Standard Meaning": meaning, "Asset Context": ctx, "Mapping Priority": prio,
             "Notes": REVIEW}
        ws.append(row_from(hdr, d))
        n += 1
    return n


# --------------------------------------------------------------------------- #
# 10 - Compatibility rules
# --------------------------------------------------------------------------- #
COMPAT = [
    {"Rule ID": "CR_CBM_01", "Rule Type": "Advisory", "Scenario ID": "TR_CBM_001", "Asset Type": "Transformer",
     "Condition / Trigger": "CBM / transformer condition-monitoring project at BOQ / quoting stage",
     "Recommended Action": (
         "Do NOT wait for or require an SLD - CBM transformer projects usually have no drawing when quoting. "
         "Drive Step 1 & 2 from the tender / project technical spec and count quantities from the monitored-parameter "
         "list (temperature points, analog inputs, RTD, CT, tap position, DGA gas count, bushings, PD sensors), "
         "normally in the 'Online Monitoring' / 'TMS-QTMS' / 'Technical Particulars' section."),
     "Severity": "Low", "Notes": REVIEW},
    {"Rule ID": "CR_CBM_02", "Rule Type": "Quantity", "Scenario ID": "TR_CBM_001", "Asset Type": "Transformer",
     "Condition / Trigger": "Sizing the QTMS modules for a transformer bank",
     "Recommended Action": (
         "Size modules by channel capacity: Analog Input 8 inputs, Digital Input 14 inputs, Fibre-Optic 4/6/8, "
         "Output Relay 8 outputs. Use a 6U chassis when module count is high, 3U otherwise."),
     "Severity": "Low", "Notes": REVIEW},
    {"Rule ID": "CR_CBM_03", "Rule Type": "Selection", "Scenario ID": "TR_DGA_001", "Asset Type": "Transformer",
     "Condition / Trigger": "DGA gas-count requirement stated in the spec",
     "Recommended Action": (
         "Select Serveron TM8 for 8-9 gas, TM3 for 3 gas, TM1 for hydrogen-only. Confirm DGA feasibility on OLTC / "
         "vacuum-interrupter tap-changer oil with the switch OEM before committing."),
     "Severity": "Medium", "Notes": REVIEW},
    {"Rule ID": "CR_CBM_04", "Rule Type": "Quantity", "Scenario ID": "TR_BUSH_001", "Asset Type": "Transformer bushing",
     "Condition / Trigger": "More than 6 bushings to be monitored on one transformer/host",
     "Recommended Action": (
         "One bushing-monitoring host handles a maximum of 6 bushings; if 9 bushings are required (e.g. 6x132kV + "
         "3x275kV), quote a second host or two bushing-monitor types."),
     "Severity": "Medium", "Notes": REVIEW + " Source: Kuwait BESS LZ263476 inquiry."},
    {"Rule ID": "CR_CBM_05", "Rule Type": "Exclusion", "Scenario ID": "TAPCHG_001", "Asset Type": "Tap changer",
     "Condition / Trigger": "Transformer has only a DETC (de-energised / off-circuit tap changer)",
     "Recommended Action": (
         "Do NOT quote OLTC online monitoring (tap position / motor current / contact wear) unless an on-load tap "
         "changer (OLTC) is present. DETC has no online tap monitoring."),
     "Severity": "Low", "Notes": REVIEW + " Source: ABB Indonesia PLN GSUT LZ250378 (DETC = N.A.)."},
]

# --------------------------------------------------------------------------- #
# 17 - Missing-info questions
# --------------------------------------------------------------------------- #
def missing(item, scen, why, question, prio="Medium", owner="Sales / Engineer"):
    return {"Project ID": "", "Missing Information Item": item, "Related Scenario ID": scen,
            "Why It Matters": why, "Suggested Customer / Engineer Question": question,
            "Priority": prio, "Owner": owner, "Status": "Open", "Notes": REVIEW}


MISSING = [
    missing("Monitored-parameter list (CBM)", "TR_CBM_001",
            "CBM quantity is counted from the parameters to be monitored, not from a drawing.",
            "Please provide the list of parameters to monitor (temperature points, analog inputs, RTD/CT, tap "
            "position, DGA gas count, bushings, PD sensors) from the tender / technical spec.", "High"),
    missing("Number of temperature / RTD points", "TR_CBM_001; TR_TEMP_001",
            "Drives the QTMS Analog Input module count.",
            "How many oil / winding / ambient / cooling / OLTC temperature (RTD) points are required?", "High"),
    missing("DGA gas count required", "TR_DGA_001",
            "Selects the Serveron model (TM1 / TM3 / TM8).",
            "How many dissolved gases must be monitored (H2 only / 3-gas / 8-9-gas)?", "High"),
    missing("Number of bushings to monitor", "TR_BUSH_001",
            "Bushing sensor count and host count (max 6 bushings per host).",
            "How many bushings per transformer must be monitored, and at which voltage levels?", "Medium"),
    missing("Tap changer type & tap-position signal", "TAPCHG_001",
            "OLTC vs DETC and 4-20mA tap-position signal availability.",
            "Is the tap changer an OLTC (on-load) or DETC (off-circuit)? Is a 4-20mA tap-position signal available?", "Medium"),
    missing("Direct winding-temperature channels", "TR_TEMP_001",
            "Sizes the fibre-optic module (4/6/8) and fibre / OFT / TWP quantities.",
            "How many direct fibre-optic winding-temperature sensors / channels are required (HV / LV / core)?", "Medium"),
]

# --------------------------------------------------------------------------- #
# 11 - SLD asset vocabulary (record-only note)
# --------------------------------------------------------------------------- #
SLD_ROWS = [
    {"Asset Type": "Transformer Bank", "Category": "Primary plant (CBM)",
     "Mapped Count Field(s)": "transformer_count",
     "Recognition Path": "Text-layer (project document)",
     "Notes": (
         "Added 2026-07 CBM KB scan. CBM transformer monitoring projects usually have NO SLD at the BOQ stage; the "
         "transformer/bank count and all monitored-parameter counts (temperature, analog input, RTD, CT, tap position, "
         "DGA gas count, bushings, PD sensors) are taken from the tender / technical-spec text "
         "(Online Monitoring / TMS-QTMS section), not from a drawing. " + REVIEW)},
]


def append_sld(ws):
    hdr, _ = header_list(ws, "Asset Type")
    have = existing_ids(ws, "Asset Type")
    n = 0
    for d in SLD_ROWS:
        if d["Asset Type"] in have:
            continue
        ws.append(row_from(hdr, d))
        n += 1
    return n


# --------------------------------------------------------------------------- #
def main():
    os.makedirs("backups", exist_ok=True)
    backup = os.path.join("backups", f"Qualitrol_BOQ_Matching_Data_Package.cbm_backup_{STAMP}.xlsx")
    shutil.copy2(XLSX, backup)
    print("backup ->", backup)

    wb = openpyxl.load_workbook(XLSX)
    counts = {}
    counts["03 scenarios (new)"] = append_dicts(wb["03_Scenario_Master"], "Scenario ID", SCENARIOS, "Scenario ID")
    counts["03 scenarios (calibrated)"] = calibrate_scenarios(wb["03_Scenario_Master"])
    counts["04 metrics"] = append_dicts(wb["04_Metric_Dictionary"], "Metric ID", METRICS, "Metric ID")
    counts["05 synonyms"] = append_synonyms(wb["05_Synonym_Mapping"])
    counts["06 families"] = append_dicts(wb["06_Product_Family_Master"], "Product Family ID", FAMILIES, "Product Family ID")
    counts["07 products (new)"] = append_dicts(wb["07_Product_Master_Template"], "Product ID", PRODUCTS, "Product ID")
    counts["07 products (calibrated)"] = calibrate_products(wb["07_Product_Master_Template"])
    counts["09 quantity_rules"] = append_dicts(wb["09_Quantity_Rules"], "Quantity Rule ID", QRULES, "Quantity Rule ID")
    counts["10 compatibility"] = append_dicts(wb["10_Compatibility_Rules"], "Rule ID", COMPAT, "Rule ID")
    counts["17 missing_info"] = append_dicts(wb["17_Missing_Info_Questions"], "Missing Information Item", MISSING, "Missing Information Item")
    counts["11 sld_vocab"] = append_sld(wb["11_SLD_Asset_Vocabulary"])

    wb.save(XLSX)
    print("done:")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
