"""Reverse-populate the Qualitrol data package from real BOQs (APPEND-ONLY).

Grounded in the real Habshan / KIZAD BOQs (775368, 776060) and the canonical
`MEA Example BOQ.xlsx`, this adds the product families, product models and
quantity rules that those engineered BOQs use but the KB was missing —
principally the shared IDM+/Informa DAUs (FMS/PQM/PMU on one device), the
panels/cabinets, network & timing gear, software/licences and services.

Safety: it ONLY appends new rows. It never edits or deletes existing rows, and
skips anything already present (dedup by Family ID / Product Model / Rule ID).
A timestamped backup of the workbook is written before saving.
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
PKG = REPO / "Qualitrol_BOQ_Matching_Data_Package.xlsx"
SRC_NOTE = "Auto-added from BOQ reverse-extraction (775368 / 776060 / MEA Example BOQ); candidate — review before use."

FAM_SHEET, PROD_SHEET, QR_SHEET, SCEN_SHEET = (
    "06_Product_Family_Master", "07_Product_Master_Template",
    "09_Quantity_Rules", "03_Scenario_Master",
)

# --- New product families (col order matches sheet header) ------------------ #
# Family ID, Family, Applicable Scenario IDs, Primary Asset Type,
# Typical Capabilities, Default Quantity Rule ID, Dependencies, Notes
FAMILIES = [
    ("PF_DAU_REC", "Multi-function DAU / Recorder (IDM+ / Informa)",
     "FMS_001;DFR_DDR_001;PMU_001;WAMS_001;PQ_CLASSA_001", "Bay / Feeder",
     "Combined fault/disturbance (FMS/DFR), power-quality (PQM) and phasor "
     "(PMU/WAMS) recording delivered on one DAU family; installed per monitored "
     "bay; channel variants 9A/32D, 18A/32D, 27A/64D.",
     "QR_DAU_BAY_001", "Monitored bay count per voltage level & function"),
    ("PF_MON_PANEL", "Monitoring Panel / Cabinet",
     "FMS_001;PMU_001;WAMS_001;PQ_CLASSA_001", "Panel",
     "FMS/PQM/WAMS field panels and LEV/PDC system cabinets housing the DAUs, "
     "industrial PC, network and GPS equipment.",
     "QR_PANEL_001", "DAU count; panels per Transco spec"),
    ("PF_NET_SEC", "Network & Security Equipment",
     "COMM_SCADA_001", "Panel",
     "Managed L2/L3 Ethernet switches (RJ45 + fiber) and firewalls for the "
     "monitoring LAN and integration to LDC/TCC.",
     "QR_PER_PANEL_001", "Panel count; comms architecture"),
    ("PF_TIMING", "GPS Timing",
     "PMU_001;WAMS_001", "System",
     "GPS antenna + cable + amplifier + surge arrestor and GPS splitters for "
     "time-synchronised PMU/WAMS recording.",
     "QR_PER_SYSTEM_001", "Time-sync requirement (PMU/WAMS present)"),
    ("PF_SW_LIC", "Monitoring Software & Licences",
     "COMM_SCADA_001;PMU_001;WAMS_001", "System",
     "iQ+ master-station software, EPG PMU/WAMS licences, and cyber software "
     "(antivirus / whitelisting / backup) — normally one set per substation.",
     "QR_PER_SYSTEM_001", "User/device tiers; licence counts"),
    ("PF_SERVICES", "Engineering & Commissioning Services",
     "COMM_SCADA_001", "System",
     "Factory FAT, testing & commissioning, communication establishment, "
     "cybersecurity, energisation assistance, training and spares (day-rate).",
     "QR_SERVICES_001", "Scope size drives service days"),
    ("PF_PANEL_ACC", "Panel Accessories",
     "FMS_001;PMU_001;PQ_CLASSA_001", "Bay",
     "Per-DAU / per-panel wiring accessories such as test switches.",
     "QR_PER_DAU_001", "DAU count"),
]

# --- New product models: (Model, Family ID, Family, Scenarios, Asset, Desc) - #
_D = "PF_DAU_REC", "Multi-function DAU / Recorder (IDM+ / Informa)"
_P = "PF_MON_PANEL", "Monitoring Panel / Cabinet"
_N = "PF_NET_SEC", "Network & Security Equipment"
_T = "PF_TIMING", "GPS Timing"
_S = "PF_SW_LIC", "Monitoring Software & Licences"
_V = "PF_SERVICES", "Engineering & Commissioning Services"
_A = "PF_PANEL_ACC", "Panel Accessories"
ALL5 = "FMS_001;DFR_DDR_001;PMU_001;WAMS_001;PQ_CLASSA_001"

PRODUCTS = [
    ("IDM+ 27A/64D", *_D, ALL5, "Bay / Feeder", "Multi-function disturbance / fault / PQ / phasor recorder DAU, 27 analogue / 64 digital channels; one per HV bay (CPU can be time master).", "QR_DAU_BAY_001"),
    ("IDM+ 18A/32D", *_D, "FMS_001;DFR_DDR_001", "Bay / Feeder", "FMS/DFR data-acquisition unit, 18 analogue / 32 digital channels.", "QR_DAU_BAY_001"),
    ("IDM+ 9A/32D", *_D, "FMS_001;PMU_001;WAMS_001", "Bay / Feeder", "Compact DAU, 9 analogue / 32 digital channels; used on 132/11kV bays and for PMU.", "QR_DAU_BAY_001"),
    ("Informa 9A/32D", *_D, "PQ_CLASSA_001", "Bay / Feeder", "Power-quality DAU (Class A), 9 analogue / 32 digital channels; one per PQ-monitored bay.", "QR_DAU_BAY_001"),
    ("EPG PMU license", *_S, "PMU_001", "System", "EPG phasor measurement (PMU) software licence, allocated per PMU channel/bay.", "QR_PER_SYSTEM_001"),
    ("EPG WAMS Software License", *_S, "WAMS_001", "System", "EPG wide-area monitoring (WAMS) software licence.", "QR_PER_SYSTEM_001"),
    ("iQ+ (5 users, 25 devices)", *_S, "COMM_SCADA_001", "System", "iQ+ master-station / fault-location software; quoted once per system (user/device tiers).", "QR_PER_SYSTEM_001"),
    ("Trend Micro Antivirus (3yr)", *_S, "COMM_SCADA_001", "System", "Endpoint antivirus for the monitoring PCs (Transco cybersecurity).", "QR_PER_SYSTEM_001"),
    ("Trellix Whitelisting Software (3yr)", *_S, "COMM_SCADA_001", "System", "Application whitelisting for the monitoring PCs.", "QR_PER_SYSTEM_001"),
    ("Acronis Backup", *_S, "COMM_SCADA_001", "System", "Backup software for the monitoring PCs.", "QR_PER_SYSTEM_001"),
    ("Industrial Rack Mounted PC", *_P, "COMM_SCADA_001", "Panel", "Industrial rack-mounted PC for FMS/PQM/PDC/PMU station (plus backup PC).", "QR_PER_SYSTEM_001"),
    ("Monitor, Keyboard", *_P, "COMM_SCADA_001", "Panel", "Operator monitor and keyboard for the LEV/PDC cabinet.", "QR_PER_SYSTEM_001"),
    ("KVM", *_P, "COMM_SCADA_001", "Panel", "KVM switch for the PDC cabinet.", "QR_PER_SYSTEM_001"),
    ("Printer", *_P, "COMM_SCADA_001", "Panel", "Report printer for the LEV cabinet.", "QR_PER_SYSTEM_001"),
    ("Alarm Annunciator", *_P, "COMM_SCADA_001", "Panel", "Alarm annunciator for the monitoring panel.", "QR_PER_SYSTEM_001"),
    ("16 Port RJ45 + fiber L2 switch", *_N, "COMM_SCADA_001", "Panel", "Managed layer-2 Ethernet switch (RJ45 + fiber uplinks) for the monitoring LAN; port/fiber count varies (4/8/16).", "QR_PER_PANEL_001"),
    ("16 Port RJ45 + fiber L3 switch", *_N, "COMM_SCADA_001", "Panel", "Managed layer-3 Ethernet switch (RJ45 + fiber) for routing to LDC/TCC.", "QR_PER_PANEL_001"),
    ("Firewall", *_N, "COMM_SCADA_001", "Panel", "Firewall for the monitoring LAN / OETC connection (optional per architecture).", "QR_PER_PANEL_001"),
    ("GPS Antenna & Cable (100m) + amplifier + mounting kit + surge arrestor", *_T, "PMU_001;WAMS_001", "System", "GPS timing antenna kit for time-synchronised PMU/WAMS.", "QR_PER_SYSTEM_001"),
    ("2-Way GPS Splitter", *_T, "PMU_001;WAMS_001", "System", "GPS signal splitter feeding multiple DAUs from one antenna.", "QR_PER_SYSTEM_001"),
    ("Test Switches", *_A, "FMS_001;PMU_001;PQ_CLASSA_001", "Bay", "Test/isolation switches per DAU/bay.", "QR_PER_DAU_001"),
    ("FMS Panel (as per Transco Spec)", *_P, "FMS_001", "Panel", "FMS field panel housing the FMS DAUs and accessories.", "QR_PANEL_001"),
    ("WAMS/PQM Panel (as per Transco Spec)", *_P, "PMU_001;WAMS_001;PQ_CLASSA_001", "Panel", "WAMS/PQM field panel housing the PMU/PQ DAUs and accessories.", "QR_PANEL_001"),
    ("LEV Cubicle (as per Transco Spec)", *_P, "COMM_SCADA_001", "Panel", "Local equipment (LEV) cubicle housing the station PC, network and software.", "QR_PANEL_001"),
    ("PDC Cubicle (as per Transco Spec)", *_P, "PMU_001;WAMS_001", "Panel", "PDC cubicle housing the PDC/PMU servers, network and software.", "QR_PANEL_001"),
    ("Factory FAT", *_V, "COMM_SCADA_001", "System", "Factory acceptance testing of the panels (day-rate; typ. 2–5 days).", "QR_SERVICES_001"),
    ("Testing & Commissioning", *_V, "COMM_SCADA_001", "System", "Site testing & commissioning (day-rate; scales with panel/bay count).", "QR_SERVICES_001"),
    ("Communication Establishment (LDC/TCC/REV-E4)", *_V, "COMM_SCADA_001", "System", "Communication link establishment with LDC/TCC/REV/E4 (day-rate).", "QR_SERVICES_001"),
    ("Cybersecurity Service", *_V, "COMM_SCADA_001", "System", "Cybersecurity hardening/config (day-rate).", "QR_SERVICES_001"),
    ("Energisation Assistance", *_V, "COMM_SCADA_001", "System", "Energisation support (day-rate + mobilisation).", "QR_SERVICES_001"),
    ("Training", *_V, "COMM_SCADA_001", "System", "Operator/maintenance training (TBA).", "QR_SERVICES_001"),
    ("Mandatory Spares 10%", *_V, "COMM_SCADA_001", "System", "Mandatory spare parts, typically 10% of hardware.", "QR_SERVICES_001"),
]

# --- New quantity rules (col order matches sheet header) -------------------- #
# Rule ID, Scenario ID, Family ID, Family, Quantity Basis, Rule Description,
# Need Drawing, Need Asset List, Count Field, Example, Assumption / Risk
QRULES = [
    ("QR_DAU_BAY_001", ALL5, "PF_DAU_REC", "Multi-function DAU / Recorder (IDM+ / Informa)",
     "Monitored bay count (per voltage level & function)",
     "One DAU per monitored bay, per function (FMS/PQM/PMU), sized by voltage "
     "level (27A/64D on 400kV OHL bays; 9A/32D on 132/11kV).",
     "Required", "Yes", "measurement_point_count",
     "14 IDM+ for 14 monitored 400kV bays; 9 Informa for 9 PQ bays.",
     "Count *monitored* bays (not raw feeder labels); confirm per-voltage split."),
    ("QR_PANEL_001", "FMS_001;PMU_001;WAMS_001;PQ_CLASSA_001", "PF_MON_PANEL", "Monitoring Panel / Cabinet",
     "Panel count from DAU grouping", "One field panel per ~2 DAUs plus one LEV/PDC system cabinet per station.",
     "Optional", "No", "panel_count", "6 FMS/WAMS panels + 5 FMS/PQM panels + LEV/PDC cabinets.",
     "Panel grouping per Transco spec; confirm with engineering."),
    ("QR_PER_PANEL_001", "COMM_SCADA_001", "PF_NET_SEC", "Network & Security Equipment",
     "Per panel", "Switches/firewalls per panel and per LEV/PDC cabinet as per comms architecture.",
     "Optional", "No", "panel_count", "2× L2 + 2× L3 switch per LEV; firewalls per cabinet.",
     "Depends on network architecture; confirm."),
    ("QR_PER_SYSTEM_001", "COMM_SCADA_001;PMU_001;WAMS_001", "PF_SW_LIC", "Monitoring Software & Licences",
     "1 per substation/system", "System-level software / licences / timing quoted once per substation "
     "(licences may scale by PMU/WAMS channel).",
     "No", "No", "site_count", "1× iQ+ per station; EPG PMU licences per PMU channel.",
     "Confirm licence/user tiers and redundancy."),
    ("QR_PER_DAU_001", "FMS_001;PMU_001;PQ_CLASSA_001", "PF_PANEL_ACC", "Panel Accessories",
     "Per DAU", "Accessories (test switches etc.) counted per DAU/bay.",
     "No", "No", "measurement_point_count", "One set of test switches per DAU.",
     "Confirm accessory list per DAU."),
    ("QR_SERVICES_001", "COMM_SCADA_001", "PF_SERVICES", "Engineering & Commissioning Services",
     "Per project (day-rate estimate)", "Service days scale with panel/bay count; quoted as day-rate + mobilisation.",
     "No", "No", "site_count", "FAT 5 days; commissioning 25 days for a large GS.",
     "Day estimates need engineer confirmation."),
]


def _header_row(rows, first_col):
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip().lower() == first_col.lower():
            return i
    raise RuntimeError(f"header {first_col!r} not found")


def main():
    wb = openpyxl.load_workbook(PKG)  # NOT read_only — we append
    # Existing scenario IDs (only reference scenarios that exist).
    sc = wb[SCEN_SHEET]; scr = list(sc.iter_rows(values_only=True))
    sh = _header_row(scr, "Scenario ID")
    scen_ids = {str(r[0]).strip() for r in scr[sh + 1:] if r and r[0]}

    def keep_scen(csv):
        keep = [s for s in csv.split(";") if s.strip() in scen_ids]
        return ";".join(keep) if keep else csv  # fall back to given if none match

    added = {"families": 0, "products": 0, "rules": 0}
    skipped = {"families": 0, "products": 0, "rules": 0}

    # Families -----------------------------------------------------------------
    fam = wb[FAM_SHEET]; fr = list(fam.iter_rows(values_only=True))
    fh = _header_row(fr, "Product Family ID")
    fam_ids = {str(r[0]).strip() for r in fr[fh + 1:] if r and r[0]}
    fam_names = {str(r[1]).strip().lower() for r in fr[fh + 1:] if r and len(r) > 1 and r[1]}
    for fid, name, scen, asset, cap, qr, dep in FAMILIES:
        if fid in fam_ids or name.strip().lower() in fam_names:
            skipped["families"] += 1; continue
        fam.append([fid, name, keep_scen(scen), asset, cap, qr, dep, SRC_NOTE])
        added["families"] += 1

    # Products -----------------------------------------------------------------
    prod = wb[PROD_SHEET]; pr = list(prod.iter_rows(values_only=True))
    ph = _header_row(pr, "Product ID")
    existing_models = {str(r[1]).strip().lower() for r in pr[ph + 1:] if r and len(r) > 1 and r[1]}
    pcount = sum(1 for r in pr[ph + 1:] if r and r[0])
    idx = 0
    for model, fid, fname, scen, asset, desc, qr in PRODUCTS:
        if model.strip().lower() in existing_models:
            skipped["products"] += 1; continue
        idx += 1
        pid = f"PROD_{fid}_{idx:02d}"
        # cols: Product ID, Model, Family ID, Family, Scenarios, Asset, Desc,
        # Standards, Protocols, Default QR, Required Acc, Optional Acc,
        # Datasheet URL, Source Owner, Status, Notes
        prod.append([pid, model, fid, fname, keep_scen(scen), asset, desc,
                     "", "", qr, "", "", "", "BOQ reverse-extraction",
                     "candidate", SRC_NOTE])
        added["products"] += 1

    # Quantity rules -----------------------------------------------------------
    qrs = wb[QR_SHEET]; qr_rows = list(qrs.iter_rows(values_only=True))
    qh = _header_row(qr_rows, "Quantity Rule ID")
    rule_ids = {str(r[0]).strip() for r in qr_rows[qh + 1:] if r and r[0]}
    for rid, scen, fid, fname, basis, desc, ndraw, nasset, cfield, ex, assum in QRULES:
        if rid in rule_ids:
            skipped["rules"] += 1; continue
        qrs.append([rid, keep_scen(scen), fid, fname, basis, desc, ndraw,
                    nasset, cfield, ex, assum])
        added["rules"] += 1

    # Backup then save ---------------------------------------------------------
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = PKG.with_name(f"Qualitrol_BOQ_Matching_Data_Package.backup_{stamp}.xlsx")
    shutil.copyfile(PKG, backup)
    wb.save(PKG)

    print("Backup written:", backup.name)
    print("ADDED  :", added)
    print("SKIPPED (already present):", skipped)
    print("Product rows before:", pcount)


if __name__ == "__main__":
    main()
