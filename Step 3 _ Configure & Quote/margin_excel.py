"""Export a Step 3 (Configure & Quote) result to a sales-safe quotation .xlsx.

Mirrors the price-list field layout but deliberately omits cost and margin data
so the file can be shared with sales / sales-ops. It shows, per configured BOQ
line: Description, Family, Qty, Unit List, Disc %, Ext List, Net (quoted) Price,
plus the project totals.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from margin_calc import compute_margins

_TITLE_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADERS = ["Description", "Family", "Qty", "Unit List", "Disc %", "Ext List", "Net Price"]
_WIDTHS = [42, 24, 8, 14, 9, 16, 16]
_MONEY = "#,##0.00"


def generate_margin_xlsx(record: dict, out_path: Path) -> Path:
    """``record`` is the saved/working Configure & Quote record (inputs + meta)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    result = compute_margins(record)
    lines = result["lines"]
    s = result["summary"]
    currency = s["currency"]
    case_ref = record.get("caseReference") or record.get("name") or "DRAFT"

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    for i, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    n = len(_HEADERS)
    last = get_column_letter(n)

    ws.merge_cells(f"A1:{last}1")
    t = ws["A1"]
    t.value = f"Qualitrol Quotation — {case_ref}  ({currency})"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = _TITLE_FILL
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    hr = 3
    for c, label in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=hr, column=c, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if c >= 3 else "left", vertical="center")

    row = hr + 1
    money_cols = {4, 6, 7}
    for ln in lines:
        vals = [
            ln["description"], ln["family"], ln["qty"], ln["unitListPrice"],
            ln["discountPct"], ln["extListPrice"], ln["extNetPrice"],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = _BORDER
            if c == 1 or c == 2:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif c in money_cols:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = _MONEY
            elif c == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '0.0"%"'
            else:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Totals row
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER
    tot = ws.cell(row=row, column=1, value="TOTALS")
    tot.font = Font(bold=True)
    ws.cell(row=row, column=6, value=s["totalListPrice"]).number_format = _MONEY
    ws.cell(row=row, column=7, value=s["totalNetPrice"]).number_format = _MONEY
    for c in (6, 7):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="right")
    row += 2

    # Quotation summary (price + discount only)
    hdr = ws.cell(row=row, column=1, value="Quotation Summary")
    hdr.font = Font(bold=True, color="1F3A5F")
    row += 1

    def _kv(label, value, *, pct=False, bold=False):
        nonlocal row
        lab = ws.cell(row=row, column=1, value=label)
        lab.alignment = Alignment(horizontal="left", indent=1)
        if bold:
            lab.font = Font(bold=True)
        val = ws.cell(row=row, column=3, value=value)
        val.alignment = Alignment(horizontal="right")
        val.number_format = '0.0"%"' if pct else _MONEY
        if bold:
            val.font = Font(bold=True)
        row += 1

    _kv("Total List Price", s["totalListPrice"])
    _kv("Total Quoted (Target) Price", s["totalNetPrice"], bold=True)
    _kv("Overall Discount", s["overallDiscountPct"], pct=True)

    ws.freeze_panes = "A4"
    wb.save(str(out_path))
    return out_path
