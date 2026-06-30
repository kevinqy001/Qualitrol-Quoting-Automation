"""Build the Step 3 product catalog from the customer price list(s).

One-off generator (run like Step 0's catalog build). Reads one or more
price-list workbooks and writes a merged, de-duplicated ``product_catalog.json``
next to this file.

Each ``*Price List`` sheet is one product family (IDM+, FL-8, INFORMA PMD-A,
Q-PMU). Within a sheet, rows in column A with no price are section sub-headers;
priced rows are individual models. Columns (consistent across sheets):

    A Description | B Part No | C Material GBP | D Material EUR | E Material USD
    F Price GBP   | G Price EUR | H Price USD   | I Margin %

Multiple workbooks are merged per family and de-duplicated by (model, part no).
Files are processed in priority order (the canonical master first), so the
master's pricing wins on a collision and the others only ADD models the master
does not already contain. Runtime is unaffected: the result is a pre-built JSON
that the web layer loads once and caches.

Usage:
    python "Step 3 _ Configure & Quote/build_catalog.py"
    python "Step 3 _ Configure & Quote/build_catalog.py" "<price list 1.xlsx>" "<price list 2.xlsx>" ...
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

_THIS_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _THIS_DIR.parent

# Canonical master first, then project price sheets that may carry extra models.
DEFAULT_PRICE_LISTS = [
    _REPO_ROOT / "Gemba Samples" / "2" / "Pricing" / "2026-05-12 IP 2026 Price List.xlsx",
    _REPO_ROOT / "Gemba Samples" / "1" / "1" / "798779" / "2. PRICING" / "Price Sheet.xlsx",
    _REPO_ROOT / "Gemba Samples" / "1" / "1" / "768938" / "2. PRICING" / "746107 Price List.xlsx",
]
CATALOG_PATH = _THIS_DIR / "product_catalog.json"

# Sheet -> (family id, display name). Only these sheets are real price lists.
FAMILY_SHEETS = {
    "IDM+  Price List": ("IDMPLUS", "IDM+ (Fault Recorder / DFR)"),
    "FL Price List": ("FL8", "FL-8 (Fault Locator)"),
    "INFORMA PMD-A Price List": ("INFORMA_PMDA", "INFORMA PMD-A (Power Quality)"),
    "Q-PMU Price List": ("QPMU", "Q-PMU (Phasor Measurement)"),
}


def _num(value):
    if value is None or value == "" or isinstance(value, str):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _find_header_row(ws) -> int:
    """Locate the row whose column B says 'Part No' (the table header)."""
    for r in range(1, min(ws.max_row, 15) + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and "part no" in b.strip().lower():
            return r
    return 7  # sensible fallback observed across the sheets


def _round(v, n=2):
    return round(v, n) if isinstance(v, (int, float)) else v


def _extract_sheet_models(ws) -> list[dict]:
    """Extract model rows from a single ``*Price List`` worksheet."""
    header = _find_header_row(ws)
    section = ""
    models: list[dict] = []
    for r in range(header + 1, ws.max_row + 1):
        desc = ws.cell(row=r, column=1).value
        if desc is None or str(desc).strip() == "":
            continue
        desc = str(desc).strip()

        part = ws.cell(row=r, column=2).value
        mat_gbp = _num(ws.cell(row=r, column=3).value)
        mat_eur = _num(ws.cell(row=r, column=4).value)
        mat_usd = _num(ws.cell(row=r, column=5).value)
        list_gbp = _num(ws.cell(row=r, column=6).value)
        list_eur = _num(ws.cell(row=r, column=7).value)
        list_usd = _num(ws.cell(row=r, column=8).value)
        margin = _num(ws.cell(row=r, column=9).value)

        has_price = any(v is not None for v in (list_gbp, list_eur, list_usd))
        has_cost = any(v is not None for v in (mat_gbp, mat_eur, mat_usd))

        # A row with text but no price/cost is a section sub-header.
        if not has_price and not has_cost:
            section = desc
            continue
        if not has_price:
            continue  # cost-only stray row; not sellable on its own

        list_price = {k: _round(v) for k, v in
                      (("GBP", list_gbp), ("EUR", list_eur), ("USD", list_usd))
                      if v is not None}
        cost = {k: _round(v) for k, v in
                (("GBP", mat_gbp), ("EUR", mat_eur), ("USD", mat_usd))
                if v is not None}

        if margin is not None:
            cat_margin = round(margin * 100, 1)
        elif list_usd and mat_usd is not None and list_usd > 0:
            cat_margin = round((1 - mat_usd / list_usd) * 100, 1)
        else:
            cat_margin = None

        models.append({
            "model": desc,
            "partNo": str(part).strip() if part not in (None, "") else "",
            "section": section,
            "listPrice": list_price,
            "cost": cost,
            "catalogMarginPct": cat_margin,
        })
    return models


def _families_from_data_package() -> list[dict]:
    """Pull the controlled monitoring product families/models from the data package.

    The price lists only cover the recorder / PQ / PMU / fault-locator product
    lines (IDM+, FL, INFORMA, Q-PMU). The monitoring scope families that Step 1/2
    actually detect — GIS PD, transformer DGA, bushing, breaker, generator,
    motor, etc. — live in ``Qualitrol_BOQ_Matching_Data_Package.xlsx`` (sheets
    06/07). We include them so every detected scenario maps to a selectable
    family/model. They carry no list price/cost (TBD in the data package), so
    those are left blank for manual entry in the calculator.
    """
    if str(_REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(_REPO_ROOT))
    try:
        from qualitrol_core.data_package import load_data_package
    except Exception as exc:  # pragma: no cover - defensive
        print(f"  (data package families unavailable: {exc})", file=sys.stderr)
        return []

    dp = load_data_package()
    products_by_fam: dict[str, list] = {}
    for p in dp.products.values():
        products_by_fam.setdefault(p.family_id, []).append(p)

    families = []
    for fid, f in dp.families.items():
        models = []
        for p in products_by_fam.get(fid, []):
            models.append({
                "model": p.model,
                "partNo": p.product_id,
                "section": getattr(f, "primary_asset_type", "") or "",
                "listPrice": {},   # TBD in the data package -> manual entry
                "cost": {},
                "catalogMarginPct": None,
            })
        families.append({
            "id": fid,
            "name": f.family_name,
            "sheet": "Data Package 06/07 (monitoring scope)",
            "priced": False,
            "modelCount": len(models),
            "models": models,
        })
    return families


def build_catalog(price_list_paths: list[Path]) -> dict:
    """Merge ``*Price List`` sheets across workbooks, de-duplicated per family,
    then append the controlled monitoring families from the data package.

    De-dup key per price-list family is (model lower, part-no lower). Files are
    processed in order; the first occurrence wins, so list earlier the workbook
    whose pricing should be authoritative.
    """
    from openpyxl import load_workbook

    # Per family: ordered dict keyed by (model, partNo) -> model record.
    fam_models: dict[str, dict[tuple[str, str], dict]] = {
        fam_id: {} for fam_id in (v[0] for v in FAMILY_SHEETS.values())
    }
    sources: list[str] = []

    for path in price_list_paths:
        if not path.exists():
            print(f"  (skip, not found) {path}", file=sys.stderr)
            continue
        sources.append(path.name)
        wb = load_workbook(str(path), data_only=True)
        for sheet_name, (fam_id, _fam_name) in FAMILY_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            for m in _extract_sheet_models(wb[sheet_name]):
                key = (m["model"].strip().lower(), m["partNo"].strip().lower())
                fam_models[fam_id].setdefault(key, m)
        wb.close()

    families = []
    for sheet_name, (fam_id, fam_name) in FAMILY_SHEETS.items():
        models = list(fam_models[fam_id].values())
        families.append({
            "id": fam_id,
            "name": fam_name,
            "sheet": sheet_name,
            "priced": True,
            "modelCount": len(models),
            "models": models,
        })

    # Append the controlled monitoring families (GIS PD, DGA, bushing, ...).
    dp_families = _families_from_data_package()
    if dp_families:
        sources.append("Qualitrol_BOQ_Matching_Data_Package.xlsx")
        families.extend(dp_families)

    return {
        "sources": sources,
        "currencies": ["GBP", "EUR", "USD"],
        "families": families,
    }


def main(argv=None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    paths = [Path(a) for a in argv] if argv else list(DEFAULT_PRICE_LISTS)
    existing = [p for p in paths if p.exists()]
    if not existing:
        print("No price list files found:\n  " + "\n  ".join(str(p) for p in paths),
              file=sys.stderr)
        return 1
    catalog = build_catalog(paths)
    CATALOG_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    total = sum(f["modelCount"] for f in catalog["families"])
    print(f"Wrote {CATALOG_PATH}")
    print(f"Sources merged: {len(catalog['sources'])}")
    for s in catalog["sources"]:
        print(f"  - {s}")
    for f in catalog["families"]:
        print(f"  {f['name']:<34} {f['modelCount']} models")
    print(f"Total: {total} models across {len(catalog['families'])} families")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
