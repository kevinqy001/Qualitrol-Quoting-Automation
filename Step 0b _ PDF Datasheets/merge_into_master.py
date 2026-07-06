"""Merge the extracted datasheet catalog into the master data package.

Does two things (steps 1 & 3 of the Step 0b plan):

  1. Promote a CURATED set of high-frequency "proposed" metrics from the
     Unmapped_Parameters list into ``04_Metric_Dictionary`` (folding obvious
     synonyms together), then re-map the matching unmapped parameters so they
     become real ``08`` rows.
  2. Merge the datasheet-sourced products / parameters into
     ``07_Product_Master_Template`` / ``08_Product_Parameter_Template``.

Conflict policy: **the new datasheet run wins**. If a product model already
exists (Tavily-sourced), its ``07`` row is overwritten in place (keeping the
existing Product ID for referential integrity) and ALL of its ``08`` parameter
rows are replaced with the datasheet-sourced ones.

Safety: the master is only overwritten after a fresh timestamped backup is made
in the same folder. Run with ``--dry-run`` to preview counts without writing.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from qualitrol_core import config  # noqa: E402

CATALOG_JSON = config.OUTPUT_DIR / "_pdf_catalog" / "step0b_pdf_catalog.json"
MANUAL_PREFIX = "Preparation/Qualitrol Product"


# --------------------------------------------------------------------------- #
# Curated metric promotion table
#   canonical id -> (Standard Name, Unit, Data Type, [normalized synonyms])
# Only clusters that appear >=3x across the datasheets are promoted; obvious
# synonyms are folded into one canonical metric to keep the dictionary clean.
# --------------------------------------------------------------------------- #
CANON_METRICS: dict[str, tuple[str, str, str, list[str]]] = {
    "MET_DIELECTRIC_STRENGTH": ("Dielectric Strength", "kV", "number",
        ["dielectric strength", "dielectric isolation"]),
    "MET_MOUNTING_TYPE": ("Mounting Type", "text", "controlled text",
        ["mounting type", "mounting", "mounting style", "mounting orientation"]),
    "MET_PROCESS_CONNECTION": ("Process / Tank Connection", "text", "text",
        ["tank connection", "tank connection thread", "mounting thread",
         "electrical connection type", "process connection"]),
    "MET_DIAL_DIAMETER": ("Dial Diameter", "mm", "number", ["dial diameter"]),
    "MET_DIAL_RANGE": ("Dial / Measurement Range", "text", "text", ["dial range"]),
    "MET_SWITCH_RATING_AC": ("Switch Contact Rating (AC)", "text", "text",
        ["switch rating ac"]),
    "MET_SWITCH_RATING_DC": ("Switch Contact Rating (DC)", "text", "text",
        ["switch rating dc", "switch dc rating"]),
    "MET_SWITCH_CONTACT_RATING": ("Switch Contact Rating", "text", "text",
        ["switch rating", "switch contact rating"]),
    "MET_NUMBER_OF_SWITCHES": ("Number of Switches", "count", "integer",
        ["number of switches"]),
    "MET_INDICATION_ACCURACY": ("Indication / Measurement Accuracy", "%", "text",
        ["indication accuracy", "measurement accuracy", "input accuracy",
         "switching accuracy", "full range accuracy", "accuracy"]),
    "MET_REPEATABILITY": ("Repeatability", "%", "text", ["repeatability"]),
    "MET_MAX_LOAD": ("Maximum Load", "text", "text", ["max load"]),
    "MET_POWER_CONSUMPTION": ("Power Consumption", "W / VA", "text",
        ["power consumption"]),
    "MET_FREQUENCY_BANDWIDTH": ("Frequency Bandwidth", "Hz", "text",
        ["frequency bandwidth", "bandwidth"]),
    "MET_FLOAT_MECHANISM": ("Float Mechanism / Material", "text", "text",
        ["float mechanism", "float material"]),
    "MET_ANALOG_OUTPUT": ("Analog Output Signal", "text", "text",
        ["analog output signal"]),
    "MET_RESPONSE_TIME": ("Response / Acquisition Time", "s", "text",
        ["response time", "data acquisition time"]),
    "MET_WEIGHT": ("Weight", "kg", "text", ["weight"]),
    "MET_PROBE_LENGTH": ("Probe / Sensor Length", "m / mm", "text",
        ["probe length"]),
    "MET_CASE_MATERIAL": ("Case / Housing Material", "text", "text",
        ["case material"]),
    "MET_OPERATING_PRESSURE": ("Operating Pressure", "bar / psi", "text",
        ["operating pressure", "oil inlet pressure"]),
    "MET_FLOW_RATE": ("Flow Rate", "text", "text", ["flow rate"]),
    "MET_PHASE_WINDOWS": ("Phase / Measurement Windows", "count", "integer",
        ["phase windows"]),
    "MET_STORAGE_TEMP": ("Storage Temperature", "°C", "text",
        ["storage temperature"]),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _build_synonym_map() -> dict[str, str]:
    out: dict[str, str] = {}
    for mid, (_n, _u, _d, syns) in CANON_METRICS.items():
        for s in syns:
            out[_norm(s)] = mid
    return out


# --------------------------------------------------------------------------- #
# Worksheet helpers (sheets = title row, description row, header row, data...)
# --------------------------------------------------------------------------- #
def _header_row_idx(ws, first_col: str) -> int:
    target = first_col.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip().lower() == target:
            return r
    raise ValueError(f"header {first_col!r} not found in {ws.title}")


def _note_rows_below(ws, header_row: int) -> int:
    """Return 1 if the row directly below the header is an inserted column-guide
    note row (blank key/ID column but other cells filled), else 0.

    This mirrors the loader's own contract: a real data row always has its key
    (first) column populated, so a row with an empty first column but text in
    other columns is definitionally the human-facing note row, never data.
    """
    r = header_row + 1
    if r > ws.max_row:
        return 0
    key = ws.cell(row=r, column=1).value
    if key is not None and str(key).strip() != "":
        return 0
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            return 1
    return 0


def _headers(ws, hrow: int) -> list[str]:
    out = []
    c = 1
    while True:
        v = ws.cell(row=hrow, column=c).value
        if v is None or str(v).strip() == "":
            break
        out.append(str(v).strip())
        c += 1
    return out


def _read_records(ws, hrow: int, headers: list[str]) -> list[dict]:
    recs = []
    start = hrow + 1 + _note_rows_below(ws, hrow)
    for r in range(start, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        recs.append({h: row[i] for i, h in enumerate(headers)})
    return recs


def _rewrite_body(ws, hrow: int, headers: list[str], records: list[dict]) -> None:
    """Delete data rows below the header and re-append `records`.

    Any inserted column-guide note row (directly below the header) is preserved.
    """
    first_data = hrow + 1 + _note_rows_below(ws, hrow)
    last = ws.max_row
    if last >= first_data:
        ws.delete_rows(first_data, last - first_data + 1)
    for rec in records:
        ws.append([rec.get(h) for h in headers])


# --------------------------------------------------------------------------- #
# Merge
# --------------------------------------------------------------------------- #
def merge(dry_run: bool = False) -> dict:
    data = json.loads(CATALOG_JSON.read_text(encoding="utf-8"))
    products = data["products"]
    mapped_params = data["product_parameters"]
    unmapped = data["unmapped_parameters"]

    syn_map = _build_synonym_map()

    # -- Step 1: promote unmapped -> canonical metrics ------------------- #
    promoted: list[dict] = []
    used_metric_ids: set[str] = set()
    for u in unmapped:
        mid = syn_map.get(_norm(u.get("proposed_metric_name", "")))
        if not mid:
            continue
        used_metric_ids.add(mid)
        row = dict(u)
        row["metric_id"] = mid
        if not row.get("unit"):
            row["unit"] = CANON_METRICS[mid][1]
        promoted.append(row)

    all_params = mapped_params + promoted

    wb = openpyxl.load_workbook(config.DATA_PACKAGE_PATH)

    # -- 04 Metric Dictionary: append promoted metrics ------------------- #
    ws4 = wb["04_Metric_Dictionary"]
    h4 = _header_row_idx(ws4, "Metric ID")
    hdr4 = _headers(ws4, h4)
    existing_mids = {
        str(ws4.cell(row=r, column=1).value).strip()
        for r in range(h4 + 1, ws4.max_row + 1)
        if ws4.cell(row=r, column=1).value
    }
    new_metric_rows = 0
    for mid in CANON_METRICS:  # stable order
        if mid not in used_metric_ids or mid in existing_mids:
            continue
        name, unit, dtype, syns = CANON_METRICS[mid]
        rec = {h: None for h in hdr4}
        rec["Metric ID"] = mid
        rec["Standard Metric Name"] = name
        rec["Synonyms / Raw Terms"] = "; ".join(syns)
        rec["Standard Unit"] = unit
        rec["Data Type"] = dtype
        rec["Applies To"] = "Various products"
        rec["Used For"] = "Product parameter matching"
        rec["Required for Matching"] = "No"
        rec["Normalization Notes"] = "Added from datasheet extraction (Step 0b)."
        if not dry_run:
            ws4.append([rec.get(h) for h in hdr4])
        new_metric_rows += 1

    # -- 07 Product Master: merge (new wins), keep existing IDs on match - #
    ws7 = wb["07_Product_Master_Template"]
    h7 = _header_row_idx(ws7, "Product ID")
    hdr7 = _headers(ws7, h7)
    recs7 = _read_records(ws7, h7, hdr7)
    by_model = {_norm(str(r.get("Product Model", ""))): r for r in recs7}
    model_to_pid: dict[str, str] = {}
    updated7 = added7 = 0

    for p in products:
        key = _norm(p["model"])
        scen = "; ".join(p.get("applicable_scenarios", []))
        src = f"{MANUAL_PREFIX}/{p.get('source_file','')}".replace("\\", "/")
        if key in by_model:
            rec = by_model[key]
            pid = str(rec.get("Product ID", "")).strip() or p["product_id"]
            updated7 += 1
        else:
            rec = {h: None for h in hdr7}
            recs7.append(rec)
            by_model[key] = rec
            pid = p["product_id"]
            added7 += 1
        rec["Product ID"] = pid
        rec["Product Model"] = p["model"]
        rec["Product Family ID"] = p.get("family_id", "")
        rec["Product Family"] = p.get("family_name", "")
        rec["Applicable Scenario IDs"] = scen
        rec["Primary Asset Type"] = p.get("primary_asset_type", "")
        rec["Product Description"] = p.get("description", "")
        rec["Supported Standards"] = p.get("supported_standards", "")
        rec["Communication Protocols"] = p.get("protocols", "")
        rec["Default Quantity Rule ID"] = p.get("default_quantity_rule_id", "")
        if "Datasheet URL" in rec:
            rec["Datasheet URL"] = src
        if "Source Owner" in rec:
            rec["Source Owner"] = "Datasheet PDF (Step 0b)"
        rec["Status"] = "Candidate"
        rec["Notes"] = f"Datasheet-sourced (Step 0b) from '{p.get('source_file','')}'; verify before quoting."
        model_to_pid[key] = pid

    if not dry_run:
        _rewrite_body(ws7, h7, hdr7, recs7)

    # -- 08 Product Parameters: replace rows for touched products -------- #
    ws8 = wb["08_Product_Parameter_Template"]
    h8 = _header_row_idx(ws8, "Product ID")
    hdr8 = _headers(ws8, h8)
    recs8 = _read_records(ws8, h8, hdr8)
    touched_pids = set(model_to_pid.values())
    kept8 = [r for r in recs8 if str(r.get("Product ID", "")).strip() not in touched_pids]

    new8: list[dict] = []
    seen8: set[tuple] = set()
    for prm in all_params:
        key = _norm(prm.get("model", ""))
        pid = model_to_pid.get(key)
        if not pid:
            continue  # parameter for a product that didn't merge (shouldn't happen)
        mid = prm.get("metric_id", "")
        pname = str(prm.get("parameter_name", "")).strip()
        dedup = (pid, mid, _norm(pname))
        if dedup in seen8:
            continue
        seen8.add(dedup)
        src = prm.get("source_file", "")
        page = prm.get("page")
        ev_src = f"{MANUAL_PREFIX}/{src}".replace("\\", "/")
        if page not in (None, ""):
            ev_src += f" (page {page})"
        rec = {h: None for h in hdr8}
        rec["Product ID"] = pid
        rec["Product Model"] = prm.get("model", "")
        rec["Product Family ID"] = prm.get("family_id", "")
        rec["Parameter ID"] = mid
        rec["Parameter Name"] = pname
        rec["Min Value"] = prm.get("min_value")
        rec["Max Value"] = prm.get("max_value")
        rec["Supported Value / Text"] = prm.get("supported_value", "")
        rec["Unit"] = prm.get("unit", "")
        if "Evidence Source / Datasheet URL" in rec:
            rec["Evidence Source / Datasheet URL"] = ev_src
        ev = str(prm.get("evidence", "")).strip()
        rec["Notes"] = (f"Datasheet-sourced (Step 0b). Evidence: {ev}"
                        if ev else "Datasheet-sourced (Step 0b); verify before quoting.")
        new8.append(rec)

    if not dry_run:
        _rewrite_body(ws8, h8, hdr8, kept8 + new8)

    summary = {
        "new_metrics_added": new_metric_rows,
        "params_promoted_via_new_metrics": len(promoted),
        "products_updated_existing": updated7,
        "products_added_new": added7,
        "param_rows_kept": len(kept8),
        "param_rows_new": len(new8),
        "param_rows_total_after": len(kept8) + len(new8),
    }

    if not dry_run:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = config.DATA_PACKAGE_PATH.with_name(
            f"{config.DATA_PACKAGE_PATH.stem}.premerge_{stamp}.xlsx"
        )
        shutil.copy2(config.DATA_PACKAGE_PATH, backup)
        wb.save(config.DATA_PACKAGE_PATH)
        summary["backup"] = str(backup)
        summary["saved"] = str(config.DATA_PACKAGE_PATH)

    return summary


def main() -> None:
    ap = argparse.ArgumentParser(description="Merge Step 0b datasheet catalog into the master data package.")
    ap.add_argument("--dry-run", action="store_true", help="preview counts; do not write")
    args = ap.parse_args()

    s = merge(dry_run=args.dry_run)
    mode = "DRY-RUN (no changes written)" if args.dry_run else "MERGE COMPLETE"
    print(f"=== {mode} ===")
    for k, v in s.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
