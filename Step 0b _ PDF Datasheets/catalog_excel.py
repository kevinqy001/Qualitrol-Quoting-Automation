"""Write the extracted datasheet catalog into a standalone candidate workbook.

Produces a NEW .xlsx (never the master data package) with sheets that mirror
07/08 plus a source index and an unmapped-parameter list, so a human can review
and paste verified rows back into ``Qualitrol_BOQ_Matching_Data_Package.xlsx``.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_PRODUCT_COLS = [
    "Product ID", "Product Model", "Product Family ID", "Product Family",
    "Applicable Scenario IDs", "Primary Asset Type", "Product Description",
    "Supported Standards", "Communication Protocols", "Default Quantity Rule ID",
    "Datasheet Source File", "Status", "Notes",
]
_PARAM_COLS = [
    "Product ID", "Product Model", "Product Family ID", "Parameter ID",
    "Parameter Name", "Min Value", "Max Value", "Supported Value / Text",
    "Unit", "Match Type", "Match Priority",
    "Evidence Source File", "Source Page", "Evidence Quote", "Notes",
]
_UNMAPPED_COLS = [
    "Product Model", "Parameter Name", "Proposed Metric Name",
    "Min Value", "Max Value", "Supported Value / Text", "Unit",
    "Evidence Source File", "Source Page", "Evidence Quote",
]
_SOURCE_COLS = [
    "Source File", "Category", "Datasheet Hash", "Pages",
    "Models Extracted", "# Models", "# Params Mapped", "# Params Unmapped", "Error",
]


def _write_header(ws, cols: list[str]) -> None:
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, cols: list[str], width: int = 22) -> None:
    for idx in range(1, len(cols) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width


def write_pdf_catalog_workbook(path: str | Path, products: list[dict],
                               parameters: list[dict], unmapped: list[dict],
                               source_index: list[dict],
                               summary: dict) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # -- 07 Product Master (candidate) ----------------------------------- #
    ws_prod = wb.active
    ws_prod.title = "07_Product_Master_Template"
    _write_header(ws_prod, _PRODUCT_COLS)
    for p in products:
        ws_prod.append([
            p.get("product_id", ""), p.get("model", ""), p.get("family_id", ""),
            p.get("family_name", ""), "; ".join(p.get("applicable_scenarios", [])),
            p.get("primary_asset_type", ""), p.get("description", ""),
            p.get("supported_standards", ""), p.get("protocols", ""),
            p.get("default_quantity_rule_id", ""), p.get("source_file", ""),
            p.get("status", "Candidate"), p.get("notes", ""),
        ])
    _autosize(ws_prod, _PRODUCT_COLS)

    # -- 08 Product Parameters (candidate, mapped) ----------------------- #
    ws_param = wb.create_sheet("08_Product_Parameter_Template")
    _write_header(ws_param, _PARAM_COLS)
    for prm in parameters:
        ws_param.append([
            prm.get("product_id", ""), prm.get("model", ""), prm.get("family_id", ""),
            prm.get("metric_id", ""), prm.get("parameter_name", ""),
            prm.get("min_value"), prm.get("max_value"),
            prm.get("supported_value", ""), prm.get("unit", ""),
            "", "",
            prm.get("source_file", ""), prm.get("page"),
            prm.get("evidence", ""),
            "Datasheet-sourced; verify before merge.",
        ])
    _autosize(ws_param, _PARAM_COLS)

    # -- Unmapped parameters (need a Metric ID decision) ----------------- #
    ws_un = wb.create_sheet("Unmapped_Parameters")
    _write_header(ws_un, _UNMAPPED_COLS)
    for u in unmapped:
        ws_un.append([
            u.get("model", ""), u.get("parameter_name", ""),
            u.get("proposed_metric_name", ""),
            u.get("min_value"), u.get("max_value"),
            u.get("supported_value", ""), u.get("unit", ""),
            u.get("source_file", ""), u.get("page"), u.get("evidence", ""),
        ])
    _autosize(ws_un, _UNMAPPED_COLS)

    # -- Source index (traceability) ------------------------------------- #
    ws_src = wb.create_sheet("Source_Index")
    _write_header(ws_src, _SOURCE_COLS)
    for s in source_index:
        ws_src.append([
            s.get("source_file", ""), s.get("category", ""), s.get("file_hash", ""),
            s.get("pages"), s.get("models", ""), s.get("n_models"),
            s.get("n_parameters"), s.get("n_unmapped"), s.get("error", ""),
        ])
    _autosize(ws_src, _SOURCE_COLS, width=26)

    # -- Run info -------------------------------------------------------- #
    ws_info = wb.create_sheet("00_Run_Info")
    _write_header(ws_info, ["Metric", "Value"])
    for k, v in summary.items():
        ws_info.append([k, v])
    ws_info.append(["source", "Datasheet PDF (Preparation/Qualitrol Product)"])
    ws_info.append(["safety", "Candidate rows only; master data package untouched."])
    _autosize(ws_info, ["Metric", "Value"], width=32)

    wb.save(path)
    return path
