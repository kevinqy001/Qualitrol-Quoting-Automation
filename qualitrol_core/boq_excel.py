"""Step 2 - finished BOQ Excel generation.

Clones the blank BOQ template (``Step 2 _ Create BOQ/BOQ_Template.xlsx``, built
from the official sample BOQ) and fills the Step 2 ``draft_boq`` lines into the
DETAILS | ITEM | QTY table, leaving the standard General / Notes boilerplate in
place. The template carries the layout, column widths and header styling so the
output matches the customary Qualitrol BOQ format.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional


def _item_text(line: dict) -> str:
    """ITEM column text = model – family description."""
    model = (line.get("product_model") or "").strip()
    desc = (line.get("product_description") or "").strip()
    return f"{model} \u2013 {desc}" if model else (desc or "TBD")


def _details_text(line: dict) -> str:
    """DETAILS column text = quantity basis, prefixed with a review flag."""
    basis = (line.get("quantity_basis") or "").strip()
    status = (line.get("review_status") or "").strip()
    prefix = f"[{status}] " if status and status.lower() != "draft" else ""
    return f"{prefix}{basis}".strip()


def _qty_text(line: dict) -> str:
    qty = line.get("quantity") or 0
    try:
        f = float(qty)
    except (TypeError, ValueError):
        return ""
    if f <= 0:
        return ""
    return str(int(f)) if f.is_integer() else f"{f:g}"


def _find_general_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().lower() == "general":
            return r
    return ws.max_row + 1


def _scenario_categories() -> dict[str, str]:
    """Map scenario_id -> top-level category (for section grouping). Best-effort."""
    try:
        from qualitrol_core.data_package import load_data_package

        dp = load_data_package()
        out: dict[str, str] = {}
        for sid, sc in dp.scenarios.items():
            cat = (sc.category or "").split("/")[0].strip()
            out[sid] = cat or "Main Scope"
        return out
    except Exception:  # noqa: BLE001
        return {}


def _grouped_boq(boq: list[dict]) -> list[tuple[str, list[dict]]]:
    """Group BOQ lines into ordered (section_name, lines) by scenario category.

    Mirrors the panel-banded layout of the manual BOQ. Falls back to a single
    "Main Scope Equipment" section when categories aren't resolvable.
    """
    cats = _scenario_categories()
    order: list[str] = []
    groups: dict[str, list[dict]] = {}
    for line in boq:
        key = cats.get(line.get("scenario_id", ""), "") or "Main Scope"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(line)
    return [(f"{k} Equipment" if k != "Main Scope" else "Main Scope Equipment",
             groups[k]) for k in order]


def generate_boq_excel(
    step2: dict,
    output_path: str | Path,
    template_path: Optional[str | Path] = None,
    project_title: Optional[str] = None,
) -> Path:
    """Generate the finished BOQ .xlsx from a Step 2 result.

    Args:
        step2: parsed ``step2_create_boq.json`` (needs ``draft_boq``).
        output_path: destination .xlsx path.
        template_path: override the blank BOQ template.
        project_title: title shown in row 1 (defaults to the project id).

    Returns:
        Path to the written workbook.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.styles.colors import Color
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    from qualitrol_core import config

    template = Path(template_path) if template_path else config.boq_template_path()
    if not template.exists():
        raise FileNotFoundError(
            f"BOQ template not found: {template}. Build it with "
            '`python "Step 2 _ Create BOQ/build_blank_boq_template.py"`.'
        )

    project_id = step2.get("project_id", "PROJECT")
    boq = step2.get("draft_boq", [])

    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]

    # Title (row 1).
    ws.cell(row=1, column=1).value = project_title or f"BOQ — {project_id}"

    # Shared styles. The section band reuses the manual BOQ's light-peach fill
    # (theme accent2, tint 0.8) so the output matches the customary look.
    band_fill = PatternFill(
        patternType="solid", fgColor=Color(theme=5, tint=0.7999816888943144)
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")
    thin_bottom = Border(bottom=Side(style="thin", color="FFBFBFBF"))

    groups = _grouped_boq(boq)
    total_rows = sum(1 + len(lines) for _, lines in groups)  # 1 band per group

    general_row = _find_general_row(ws)
    if total_rows:
        ws.insert_rows(general_row, amount=total_rows)
        r = general_row
        for section_name, lines in groups:
            # Colored section band: "Desc.: <Section>" | "DEVICE SUMMARY".
            ba = ws.cell(row=r, column=1, value=f"Desc.: {section_name}")
            bb = ws.cell(row=r, column=2, value="DEVICE SUMMARY")
            bc = ws.cell(row=r, column=3, value="")
            for cell in (ba, bb, bc):
                cell.fill = band_fill
            ba.font = Font(bold=True)
            bb.font = Font(bold=True)
            r += 1
            # Equipment lines for this section.
            for j, line in enumerate(lines):
                a = ws.cell(row=r, column=1, value=_details_text(line))
                b = ws.cell(row=r, column=2, value=_item_text(line))
                c = ws.cell(row=r, column=3, value=_qty_text(line))
                a.alignment = wrap
                b.alignment = wrap
                b.font = Font(bold=True)
                c.alignment = center
                c.font = Font(bold=True)
                if j == len(lines) - 1:  # separator under the last line
                    for cell in (a, b, c):
                        cell.border = thin_bottom
                r += 1

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
