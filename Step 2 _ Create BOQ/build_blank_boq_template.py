"""One-off builder: create the blank Step 2 BOQ Excel template.

Derives ``BOQ_Template.xlsx`` (saved next to this script) from the official
sample BOQ ``Gemba Samples/1/1/773306/1. SPECS/Updated BOQ.xlsx`` by stripping
all project-specific content while preserving the layout, column widths, header
styling and the standard General / Notes boilerplate.

Result layout (sheet "BOQ"):
    Row 1  : [Project Title]              (merged A1:C1, bold)
    Row 3  : DETAILS | ITEM | QTY         (bold header)
    Row 4+ : <data insertion point>       (filled per run by boq_excel.py)
    ...    : General                      (standard scope items)
    ...    : Notes to contractor          (standard contractor notes)

Run once (or re-run if the source layout changes):
    python "Step 2 _ Create BOQ/build_blank_boq_template.py"
"""

from __future__ import annotations

import sys
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import openpyxl  # noqa: E402

SOURCE = (_REPO_ROOT / "Gemba Samples" / "1" / "1" / "773306" / "1. SPECS"
          / "Updated BOQ.xlsx")
OUTPUT = Path(__file__).resolve().parent / "BOQ_Template.xlsx"

TITLE_PLACEHOLDER = "[Project Title]"

# Project-specific tokens in the General / Notes boilerplate -> neutralised.
_NEUTRALISE = [
    ("( 12 if 33kV Panel is taken)", ""),
    ("at Ibri OETC S/S", "at site"),
    ("Ibri OETC", "[Site]"),
    ("IBRI", ""),
    ("Ibri", ""),
]


def _neutralise(text: str) -> str:
    for old, new in _NEUTRALISE:
        text = text.replace(old, new)
    return " ".join(text.split())  # collapse leftover double spaces


def build() -> Path:
    if not SOURCE.exists():
        raise FileNotFoundError(f"Source BOQ not found: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE)
    ws = wb[wb.sheetnames[0]]
    ws.title = "BOQ"

    # Locate the "General" section (start of the boilerplate tail).
    general_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip().lower() == "general":
            general_row = r
            break
    if general_row is None:
        general_row = ws.max_row + 1

    # Unmerge ALL ranges up front (while their cells still exist) so deleting
    # rows can't leave a dangling merge; the title merge is restored afterwards.
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    # Delete the project-specific data rows (panels / feeders / items), leaving
    # the header (row 3) directly above the General section.
    n_delete = general_row - 1 - 4 + 1
    if n_delete > 0:
        ws.delete_rows(4, n_delete)

    # Restore only the title merge.
    ws.merge_cells("A1:C1")

    # Title -> placeholder.
    ws["A1"] = TITLE_PLACEHOLDER

    # Neutralise project-specific text in the remaining (General / Notes) rows.
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and cell.value.strip():
                new = _neutralise(cell.value)
                if new != cell.value:
                    cell.value = new

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    out = build()
    print(f"Blank BOQ template written to: {out}")
